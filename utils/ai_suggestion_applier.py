from __future__ import annotations

import json
import re
import shutil
from calendar import monthrange
from copy import copy
from datetime import datetime, date
from pathlib import Path
from typing import Any, cast

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


# =========================
# 基础工具函数
# =========================

def _strip_code_fence(text: str) -> str:
    """
    兼容 LLM 输出 ```json ... ``` 的情况。
    """
    text = (text or "").strip()

    if text.startswith("```"):
        text = re.sub(r"^```(?:json)?", "", text, flags=re.IGNORECASE).strip()
        text = re.sub(r"```$", "", text).strip()

    return text


def _to_float(value: Any) -> float:
    if value is None:
        return 0.0

    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).replace(",", "").strip()
    if text == "":
        return 0.0

    return float(text)


def _clean_code(value: Any) -> str:
    """
    统一清洗科目编码，避免 Excel 把 400101 读成 400101.0。
    """
    if value is None:
        return ""

    text = str(value).strip()
    if text.endswith(".0"):
        text = text[:-2]

    return text


def _normalize_header(value: Any) -> str:
    """
    表头归一化：去空格、星号、换行、引号，便于兼容模板表头。
    """
    if value is None:
        return ""

    text = str(value).strip()
    text = text.replace(" ", "").replace("\n", "").replace("\r", "")
    text = text.replace("*", "").replace('"', "").replace("：", ":")
    return text


def _normalize_date(value: Any) -> str:
    """
    将日期统一为 YYYY-MM-DD。
    """
    if value is None:
        raise ValueError("交易日期为空")

    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")

    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")

    text = str(value).strip()
    if not text:
        raise ValueError("交易日期为空")

    text = text.replace("/", "-")

    if " " in text:
        text = text.split(" ")[0]

    parts = text.split("-")
    if len(parts) != 3:
        raise ValueError(f"无法识别交易日期：{value}")

    year = int(float(parts[0]))
    month = int(float(parts[1]))
    day = int(float(parts[2]))

    return f"{year:04d}-{month:02d}-{day:02d}"


def _month_end(date_text: str) -> str:
    year, month, _ = [int(x) for x in date_text.split("-")]
    last_day = monthrange(year, month)[1]
    return f"{year:04d}-{month:02d}-{last_day:02d}"


def _month_label(date_text: str) -> str:
    """
    2026-03-17 -> 3月
    """
    _, month, _ = [int(x) for x in date_text.split("-")]
    return f"{month}月"


def _is_truthy(value: Any) -> bool:
    """
    兼容 true / "true" / "True" / 1。
    """
    if isinstance(value, bool):
        return value

    if isinstance(value, (int, float)):
        return value != 0

    text = str(value).strip().lower()
    return text in {"true", "1", "yes", "y", "是"}


def _parse_ai_suggestions(ai_suggestions_json: str) -> list[dict[str, Any]]:
    """
    解析 LLM1 输出的 JSON。
    支持：
    1. {"suggestions": [...]}
    2. [...]
    """
    raw = _strip_code_fence(ai_suggestions_json)

    if not raw:
        return []

    data = json.loads(raw)

    if isinstance(data, list):
        suggestions = data
    elif isinstance(data, dict):
        suggestions = data.get("suggestions", []) or []
    else:
        suggestions = []

    if not isinstance(suggestions, list):
        return []

    return [item for item in suggestions if isinstance(item, dict)]


def _load_valid_subject_codes(subject_file: Path) -> set[str]:
    """
    从会计科目表读取合法科目编码。
    防止 LLM 编造会计科目。
    """
    if not subject_file.exists():
        return set()

    df = pd.read_excel(subject_file)

    code_col = None
    for col in df.columns:
        col_text = str(col)
        if "科目编码" in col_text or col_text.strip() in {"编码", "科目"}:
            code_col = col
            break

    if code_col is None:
        for col in df.columns:
            if "编码" in str(col):
                code_col = col
                break

    if code_col is None:
        return set()

    codes: set[str] = set()
    for value in df[code_col].dropna().tolist():
        code = _clean_code(value)
        if code:
            codes.add(code)

    return codes


def _copy_cell_style(src_cell, dst_cell) -> None:
    """
    复制单元格样式。
    """
    if src_cell.has_style:
        dst_cell._style = copy(src_cell._style)

    if src_cell.number_format:
        dst_cell.number_format = src_cell.number_format

    if src_cell.font:
        dst_cell.font = copy(src_cell.font)

    if src_cell.fill:
        dst_cell.fill = copy(src_cell.fill)

    if src_cell.border:
        dst_cell.border = copy(src_cell.border)

    if src_cell.alignment:
        dst_cell.alignment = copy(src_cell.alignment)

    if src_cell.protection:
        dst_cell.protection = copy(src_cell.protection)


def _copy_row_style(ws: Worksheet, source_row: int, target_row: int) -> None:
    """
    复制整行样式。
    """
    if source_row < 1 or source_row > ws.max_row:
        return

    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height

    for col in range(1, ws.max_column + 1):
        src = ws.cell(row=source_row, column=col)
        dst = ws.cell(row=target_row, column=col)
        _copy_cell_style(src, dst)


# =========================
# 表头识别与读写工具
# =========================

def _detect_header_row(ws: Worksheet, expected_headers: list[str], scan_rows: int = 10) -> int:
    """
    自动识别表头行。
    凭证模板第 1 行可能是导入须知，第 2 行才是真正字段表头，所以不能写死第 1 行。
    """
    expected = {_normalize_header(x) for x in expected_headers}

    best_row = 1
    best_score = -1

    max_scan = min(ws.max_row, scan_rows)

    for row in range(1, max_scan + 1):
        values = {
            _normalize_header(ws.cell(row=row, column=col).value)
            for col in range(1, ws.max_column + 1)
        }
        score = len(expected & values)

        if score > best_score:
            best_score = score
            best_row = row

    return best_row


def _get_header_map(ws: Worksheet, header_row: int) -> dict[str, int]:
    """
    读取表头行，返回归一化列名 -> 列号。
    """
    header_map: dict[str, int] = {}

    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=header_row, column=col).value
        key = _normalize_header(value)

        if key:
            header_map[key] = col

    return header_map


def _col(header_map: dict[str, int], *names: str) -> int | None:
    """
    根据候选字段名找列号。
    """
    for name in names:
        key = _normalize_header(name)
        if key in header_map:
            return header_map[key]

    return None


def _set_by_names(
    ws: Worksheet,
    row: int,
    header_map: dict[str, int],
    names: list[str],
    value: Any
) -> None:
    col = _col(header_map, *names)
    if col is not None:
        ws.cell(row=row, column=col).value = value


def _get_by_names(
    ws: Worksheet,
    row: int,
    header_map: dict[str, int],
    names: list[str],
) -> Any:
    col = _col(header_map, *names)
    if col is None:
        return None

    return ws.cell(row=row, column=col).value


# =========================
# 摘要处理
# =========================

def _normalize_ai_summary(suggestion: dict[str, Any]) -> str:
    """
    统一 AI 补录摘要。
    对 400101 收入兜底建议，优先生成“划来X月代收党费”的表达。
    """
    direction = str(suggestion.get("direction", "")).strip()
    subject_code = _clean_code(suggestion.get("candidate_subject_code"))
    transaction_date = _normalize_date(suggestion.get("transaction_date"))
    month_text = _month_label(transaction_date)

    counterparty = str(suggestion.get("counterparty", "")).strip()
    summary = str(suggestion.get("suggested_summary", "")).strip()

    if direction == "收入" and subject_code == "400101":
        if counterparty:
            return f"收到{counterparty}划来{month_text}代收党费"

    if summary:
        return summary

    if direction == "收入":
        return f"收到{counterparty}划来党费"

    return f"支付{counterparty}相关党费业务款项"


# =========================
# 凭证处理
# =========================

def _voucher_header_info(ws: Worksheet) -> tuple[int, dict[str, int]]:
    header_row = _detect_header_row(
        ws,
        expected_headers=[
            "* 核算账簿",
            "* 凭证类别",
            "* 制单日期",
            "* 摘要",
            "表头自定义项3",
            "* 科目编码",
            "* 原币借方金额",
            "* 本币借方金额",
            "* 原币贷方金额",
            "* 本币贷方金额",
            "结算日期",
            "核销业务日期",
        ],
        scan_rows=8,
    )
    return header_row, _get_header_map(ws, header_row)


def _voucher_date_for_row(ws: Worksheet, row: int, header_map: dict[str, int]) -> str | None:
    """
    优先用结算日期排序；没有结算日期时用制单日期。
    """
    value = _get_by_names(ws, row, header_map, ["结算日期", "checkdate"])
    if value is None:
        value = _get_by_names(ws, row, header_map, ["* 制单日期", "制单日期", "main_prepareddate"])

    if value is None:
        return None

    try:
        return _normalize_date(value)
    except Exception:
        return None


def _find_voucher_insert_row(
    ws: Worksheet,
    header_row: int,
    header_map: dict[str, int],
    target_transaction_date: str,
) -> int:
    """
    按结算日期找到凭证插入位置。
    规则：插到第一条结算日期大于目标日期的凭证行之前。
    """
    data_start = header_row + 1
    target = _normalize_date(target_transaction_date)

    for row in range(data_start, ws.max_row + 1):
        current = _voucher_date_for_row(ws, row, header_map)
        if current is None:
            continue

        if current > target:
            return row

    return ws.max_row + 1


def _write_voucher_pair(
    ws: Worksheet,
    row1: int,
    row2: int,
    suggestion: dict[str, Any],
    header_map: dict[str, int],
    maker: str,
    book_code: str,
    voucher_type: str,
) -> None:
    """
    写入一笔 AI 补录凭证的借贷两行。
    """
    direction = str(suggestion.get("direction", "")).strip()
    transaction_date = _normalize_date(suggestion.get("transaction_date"))
    voucher_date = _month_end(transaction_date)

    amount = _to_float(suggestion.get("amount"))
    subject_code = _clean_code(suggestion.get("candidate_subject_code"))
    summary = _normalize_ai_summary(suggestion)

    common_values = [
        (["* 核算账簿", "核算账簿"], book_code),
        (["* 凭证类别", "凭证类别"], voucher_type),
        (["* 凭证号", "凭证号"], None),
        (["附单据数"], None),
        (["* 制单人", "制单人"], maker),
        (["* 制单日期", "制单日期"], voucher_date),
        (["审核人"], None),
        (["审核日期"], None),
        (["* 摘要", "摘要"], summary),
        (["表头自定义项2"], "AI建议补录-待复核"),
        (["* 币种", "币种"], "CNY"),
        (["结算号"], None),
        (["结算日期"], transaction_date),
        (["结算方式"], None),
        (["核销号"], None),
        (["核销业务日期"], transaction_date),
    ]

    for row in [row1, row2]:
        for names, value in common_values:
            _set_by_names(ws, row, header_map, names, value)

    if direction == "收入":
        # 借：银行存款 1002
        _set_by_names(ws, row1, header_map, ["* 科目编码", "科目编码"], "1002")
        _set_by_names(ws, row1, header_map, ["* 原币借方金额", "原币借方金额"], amount)
        _set_by_names(ws, row1, header_map, ["* 本币借方金额", "本币借方金额"], amount)
        _set_by_names(ws, row1, header_map, ["* 原币贷方金额", "原币贷方金额"], None)
        _set_by_names(ws, row1, header_map, ["* 本币贷方金额", "本币贷方金额"], None)

        # 贷：候选收入科目
        _set_by_names(ws, row2, header_map, ["* 科目编码", "科目编码"], subject_code)
        _set_by_names(ws, row2, header_map, ["* 原币借方金额", "原币借方金额"], None)
        _set_by_names(ws, row2, header_map, ["* 本币借方金额", "本币借方金额"], None)
        _set_by_names(ws, row2, header_map, ["* 原币贷方金额", "原币贷方金额"], amount)
        _set_by_names(ws, row2, header_map, ["* 本币贷方金额", "本币贷方金额"], amount)

    elif direction == "支出":
        # 借：候选支出科目
        _set_by_names(ws, row1, header_map, ["* 科目编码", "科目编码"], subject_code)
        _set_by_names(ws, row1, header_map, ["* 原币借方金额", "原币借方金额"], amount)
        _set_by_names(ws, row1, header_map, ["* 本币借方金额", "本币借方金额"], amount)
        _set_by_names(ws, row1, header_map, ["* 原币贷方金额", "原币贷方金额"], None)
        _set_by_names(ws, row1, header_map, ["* 本币贷方金额", "本币贷方金额"], None)

        # 贷：银行存款 1002
        _set_by_names(ws, row2, header_map, ["* 科目编码", "科目编码"], "1002")
        _set_by_names(ws, row2, header_map, ["* 原币借方金额", "原币借方金额"], None)
        _set_by_names(ws, row2, header_map, ["* 本币借方金额", "本币借方金额"], None)
        _set_by_names(ws, row2, header_map, ["* 原币贷方金额", "原币贷方金额"], amount)
        _set_by_names(ws, row2, header_map, ["* 本币贷方金额", "本币贷方金额"], amount)

    else:
        raise ValueError(f"无法识别 AI 建议方向：{direction}")


def _insert_voucher_pair(
    ws: Worksheet,
    header_row: int,
    header_map: dict[str, int],
    suggestion: dict[str, Any],
    maker: str,
    book_code: str,
    voucher_type: str,
) -> None:
    """
    按结算日期插入凭证借贷两行，并复制相邻行格式。
    """
    transaction_date = _normalize_date(suggestion.get("transaction_date"))
    insert_row = _find_voucher_insert_row(ws, header_row, header_map, transaction_date)

    ws.insert_rows(insert_row, amount=2)

    if insert_row + 2 <= ws.max_row:
        source_row = insert_row + 2
    else:
        source_row = max(header_row + 1, insert_row - 1)

    _copy_row_style(ws, source_row, insert_row)
    _copy_row_style(ws, source_row, insert_row + 1)

    _write_voucher_pair(
        ws=ws,
        row1=insert_row,
        row2=insert_row + 1,
        suggestion=suggestion,
        header_map=header_map,
        maker=maker,
        book_code=book_code,
        voucher_type=voucher_type,
    )


def _renumber_voucher_custom_no(ws: Worksheet, header_row: int, header_map: dict[str, int]) -> None:
    """
    表头自定义项3按月重新编号。
    借贷两行共用同一个编号。
    """
    custom_col = _col(header_map, "表头自定义项3")
    summary_col = _col(header_map, "* 摘要", "摘要")

    if custom_col is None:
        return

    counters: dict[str, int] = {}
    row = header_row + 1

    while row <= ws.max_row:
        voucher_date = _get_by_names(ws, row, header_map, ["* 制单日期", "制单日期"])
        if voucher_date is None:
            row += 1
            continue

        try:
            voucher_date_text = _normalize_date(voucher_date)
        except Exception:
            row += 1
            continue

        month_key = voucher_date_text[:7]
        counters[month_key] = counters.get(month_key, 0) + 1
        current_no = counters[month_key]

        ws.cell(row=row, column=custom_col).value = current_no

        if row + 1 <= ws.max_row:
            same_business = True

            if summary_col is not None:
                cur_summary = ws.cell(row=row, column=summary_col).value
                next_summary = ws.cell(row=row + 1, column=summary_col).value
                same_business = cur_summary == next_summary

            cur_date = _get_by_names(ws, row, header_map, ["结算日期", "* 制单日期", "制单日期"])
            next_date = _get_by_names(ws, row + 1, header_map, ["结算日期", "* 制单日期", "制单日期"])

            if same_business and cur_date == next_date:
                ws.cell(row=row + 1, column=custom_col).value = current_no
                row += 2
                continue

        row += 1


def _get_voucher_summary_order(ws: Worksheet, header_row: int, header_map: dict[str, int]) -> dict[str, int]:
    """
    从最终凭证中读取摘要顺序，用于台账排序。
    每个业务的借贷两行摘要相同，只记录第一次出现。
    """
    summary_col = _col(header_map, "* 摘要", "摘要")
    order: dict[str, int] = {}

    if summary_col is None:
        return order

    idx = 0
    for row in range(header_row + 1, ws.max_row + 1):
        summary = ws.cell(row=row, column=summary_col).value
        if summary is None:
            continue

        summary_text = str(summary).strip()
        if not summary_text:
            continue

        if summary_text not in order:
            order[summary_text] = idx
            idx += 1

    return order


# =========================
# 台账处理
# =========================

def _ledger_header_info(ws: Worksheet) -> tuple[int, dict[str, int]]:
    header_row = _detect_header_row(
        ws,
        expected_headers=["年", "月", "日", "摘要", "支出", "收入", "标签"],
        scan_rows=5,
    )
    return header_row, _get_header_map(ws, header_row)


def _find_ledger_total_row(ws: Worksheet, header_row: int, header_map: dict[str, int]) -> int:
    """
    查找台账明细合计行。
    """
    summary_col = _col(header_map, "摘要")

    if summary_col is None:
        return ws.max_row + 1

    for row in range(header_row + 1, ws.max_row + 1):
        value = ws.cell(row=row, column=summary_col).value
        if value is None:
            continue

        text = str(value).strip()
        if text == "合计":
            return row

    return ws.max_row + 1


def _write_ledger_row(
    ws: Worksheet,
    row: int,
    suggestion: dict[str, Any],
    header_map: dict[str, int],
) -> None:
    direction = str(suggestion.get("direction", "")).strip()
    transaction_date = _normalize_date(suggestion.get("transaction_date"))
    voucher_date = _month_end(transaction_date)

    year, month, day = [int(x) for x in voucher_date.split("-")]
    amount = _to_float(suggestion.get("amount"))

    summary = _normalize_ai_summary(suggestion)
    ledger_tag = str(suggestion.get("ledger_tag", "")).strip()

    if not ledger_tag:
        ledger_tag = "支出" if direction == "支出" else "公司党员统一上缴"

    _set_by_names(ws, row, header_map, ["年"], year)
    _set_by_names(ws, row, header_map, ["月"], f"{month:02d}")
    _set_by_names(ws, row, header_map, ["日"], f"{day:02d}")
    _set_by_names(ws, row, header_map, ["编号"], "")
    _set_by_names(ws, row, header_map, ["摘要"], summary)
    _set_by_names(ws, row, header_map, ["余额（元）", "余额"], "")
    _set_by_names(ws, row, header_map, ["标签"], ledger_tag)

    # 关键：收入行必须清空支出列，支出行必须清空收入列。
    # 先显式清空收入/支出两列，防止复制行或重排行时残留旧金额。
    _set_by_names(ws, row, header_map, ["支出"], "")
    _set_by_names(ws, row, header_map, ["收入"], "")

    if direction == "收入":
        _set_by_names(ws, row, header_map, ["收入"], amount)
    elif direction == "支出":
        _set_by_names(ws, row, header_map, ["支出"], amount)
    else:
        raise ValueError(f"无法识别 AI 建议方向：{direction}")


def _insert_ledger_rows_before_total(
    ws: Worksheet,
    header_row: int,
    header_map: dict[str, int],
    suggestions: list[dict[str, Any]],
) -> None:
    """
    先把 AI 补录台账行插到“合计”行之前，后续再按凭证摘要顺序重排。
    """
    if not suggestions:
        return

    total_row = _find_ledger_total_row(ws, header_row, header_map)

    if total_row > ws.max_row:
        total_row = ws.max_row + 1

    ws.insert_rows(total_row, amount=len(suggestions))

    source_row = max(header_row + 1, total_row - 1)

    for offset, suggestion in enumerate(suggestions):
        target_row = total_row + offset
        _copy_row_style(ws, source_row, target_row)

        # 插入行先清空所有值，防止模板或移动后的旧值残留。
        for col in range(1, ws.max_column + 1):
            ws.cell(row=target_row, column=col).value = None

        _write_ledger_row(ws, target_row, suggestion, header_map)


def _capture_row(ws: Worksheet, row: int) -> dict[str, Any]:
    values = []
    styles = []

    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        values.append(cell.value)
        styles.append({
            "style": copy(cell._style),
            "number_format": cell.number_format,
            "font": copy(cell.font),
            "fill": copy(cell.fill),
            "border": copy(cell.border),
            "alignment": copy(cell.alignment),
            "protection": copy(cell.protection),
        })

    return {
        "values": values,
        "styles": styles,
        "height": ws.row_dimensions[row].height,
    }


def _restore_row(ws: Worksheet, row: int, row_data: dict[str, Any]) -> None:
    """
    恢复行数据和样式。

    注意：
    openpyxl 的 ws.cell(row, col, value=None) 不会清空原单元格值，
    必须显式 cell.value = value。否则收入/支出列会残留旧金额，
    导致“收入行旁边出现支出金额”或“支出行旁边出现收入金额”。
    """
    ws.row_dimensions[row].height = row_data.get("height")

    values = row_data["values"]
    styles = row_data["styles"]

    for col, value in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = value

        style = styles[col - 1]
        cell._style = copy(style["style"])
        cell.number_format = style["number_format"]
        cell.font = copy(style["font"])
        cell.fill = copy(style["fill"])
        cell.border = copy(style["border"])
        cell.alignment = copy(style["alignment"])
        cell.protection = copy(style["protection"])


def _reorder_ledger_by_voucher_order(
    ws: Worksheet,
    header_row: int,
    header_map: dict[str, int],
    voucher_summary_order: dict[str, int],
) -> None:
    """
    让台账明细顺序与凭证业务顺序一致。
    """
    summary_col = _col(header_map, "摘要")
    if summary_col is None:
        return

    total_row = _find_ledger_total_row(ws, header_row, header_map)
    if total_row <= header_row + 1:
        return

    detail_start = header_row + 1
    detail_end = total_row - 1

    captured_rows = []

    for row in range(detail_start, detail_end + 1):
        row_data = _capture_row(ws, row)
        summary = ws.cell(row=row, column=summary_col).value
        summary_text = "" if summary is None else str(summary).strip()

        original_index = row - detail_start
        sort_index = voucher_summary_order.get(summary_text, 100000 + original_index)

        captured_rows.append((sort_index, original_index, row_data))

    captured_rows.sort(key=lambda x: (x[0], x[1]))

    for offset, (_, __, row_data) in enumerate(captured_rows):
        _restore_row(ws, detail_start + offset, row_data)


def _clear_wrong_direction_amounts(ws: Worksheet, header_row: int, header_map: dict[str, int]) -> None:
    """
    兜底清理：每条台账明细只能有收入或支出一个方向的金额。
    如果标签为“支出”，清空收入列；如果存在收入金额且标签不是支出，清空支出列。
    """
    expense_col = _col(header_map, "支出")
    income_col = _col(header_map, "收入")
    tag_col = _col(header_map, "标签")
    summary_col = _col(header_map, "摘要")

    if expense_col is None or income_col is None:
        return

    total_row = _find_ledger_total_row(ws, header_row, header_map)
    detail_end = total_row - 1 if total_row <= ws.max_row else ws.max_row

    for row in range(header_row + 1, detail_end + 1):
        tag = ""
        if tag_col is not None:
            tag = str(ws.cell(row=row, column=tag_col).value or "").strip()

        summary = ""
        if summary_col is not None:
            summary = str(ws.cell(row=row, column=summary_col).value or "").strip()

        expense = _to_float(ws.cell(row=row, column=expense_col).value)
        income = _to_float(ws.cell(row=row, column=income_col).value)

        if tag == "支出":
            ws.cell(row=row, column=income_col).value = ""
            continue

        if income > 0:
            ws.cell(row=row, column=expense_col).value = ""
            continue

        if expense > 0 and ("支付" in summary or tag == "支出"):
            ws.cell(row=row, column=income_col).value = ""


def _recalc_ledger_total_and_income_composition(
    ws: Worksheet,
    header_row: int,
    header_map: dict[str, int],
) -> None:
    """
    重算台账合计和收入组成。
    """
    summary_col = _col(header_map, "摘要")
    expense_col = _col(header_map, "支出")
    income_col = _col(header_map, "收入")
    tag_col = _col(header_map, "标签")

    if summary_col is None or income_col is None or expense_col is None:
        return

    _clear_wrong_direction_amounts(ws, header_row, header_map)

    total_row = _find_ledger_total_row(ws, header_row, header_map)
    if total_row > ws.max_row:
        return

    detail_start = header_row + 1
    detail_end = total_row - 1

    total_income = 0.0
    total_expense = 0.0

    income_by_tag = {
        "公司党员统一上缴": 0.0,
        "其他党员自行上缴": 0.0,
        "广发下拨": 0.0,
        "党费账户利息收入": 0.0,
    }

    for row in range(detail_start, detail_end + 1):
        income = _to_float(ws.cell(row=row, column=income_col).value)
        expense = _to_float(ws.cell(row=row, column=expense_col).value)

        total_income += income
        total_expense += expense

        tag = ""
        if tag_col is not None:
            tag_value = ws.cell(row=row, column=tag_col).value
            tag = "" if tag_value is None else str(tag_value).strip()

        if income > 0 and tag in income_by_tag:
            income_by_tag[tag] += income

    ws.cell(row=total_row, column=summary_col).value = "合计"
    ws.cell(row=total_row, column=expense_col).value = round(total_expense, 2)
    ws.cell(row=total_row, column=income_col).value = round(total_income, 2)

    label_values = {
        "公司党员统一上缴": round(income_by_tag["公司党员统一上缴"], 2),
        "其他党员自行上缴": round(income_by_tag["其他党员自行上缴"], 2),
        "广发下拨": round(income_by_tag["广发下拨"], 2),
        "党费账户利息收入": round(income_by_tag["党费账户利息收入"], 2),
        "合计：": round(sum(income_by_tag.values()), 2),
        "合计": round(sum(income_by_tag.values()), 2),
    }

    for row in range(total_row + 1, ws.max_row + 1):
        label = ws.cell(row=row, column=summary_col).value
        if label is None:
            continue

        label_text = str(label).strip()

        if label_text in label_values:
            ws.cell(row=row, column=income_col).value = label_values[label_text]
            ws.cell(row=row, column=expense_col).value = ""


# =========================
# AI 建议校验
# =========================

def _validate_suggestion(
    suggestion: dict[str, Any],
    valid_subject_codes: set[str],
) -> tuple[bool, str]:
    subject_code = _clean_code(suggestion.get("candidate_subject_code"))
    direction = str(suggestion.get("direction", "")).strip()
    amount = _to_float(suggestion.get("amount"))

    if direction not in {"收入", "支出"}:
        return False, f"方向不是收入/支出：{direction}"

    if amount <= 0:
        return False, f"金额必须大于 0：{amount}"

    if not subject_code:
        return False, "候选科目编码为空"

    if subject_code == "1002":
        return False, "候选科目不能是银行存款 1002"

    if valid_subject_codes and subject_code not in valid_subject_codes:
        return False, f"候选科目编码不在会计科目表中：{subject_code}"

    try:
        _normalize_date(suggestion.get("transaction_date"))
    except Exception as exc:
        return False, str(exc)

    return True, ""


# =========================
# 主入口
# =========================

def apply_ai_suggestions_to_drafts(
    voucher_file_path: Path,
    ledger_file_path: Path,
    output_dir: Path,
    run_id: str,
    ai_suggestions_json: str,
    subject_file: Path,
    maker: str,
    book_code: str,
    voucher_type: str,
) -> dict[str, Any]:
    """
    根据 LLM1 输出的 AI 补录建议，生成“AI补录版”凭证草稿和台账草稿。

    关键规则：
    1. 不覆盖原始草稿；
    2. 候选科目必须存在于会计科目表；
    3. 凭证按结算日期插入，不追加到底部；
    4. 台账按最终凭证顺序排列；
    5. 台账合计和收入组成重新计算；
    6. AI 补录内容标记为待人工复核；
    7. 台账每一行只允许收入或支出一个方向的金额。
    """
    output_dir.mkdir(parents=True, exist_ok=True)

    if not voucher_file_path.exists():
        raise FileNotFoundError(f"凭证草稿文件不存在：{voucher_file_path}")

    if not ledger_file_path.exists():
        raise FileNotFoundError(f"台账草稿文件不存在：{ledger_file_path}")

    suggestions = _parse_ai_suggestions(ai_suggestions_json)
    valid_subject_codes = _load_valid_subject_codes(subject_file)

    modified_voucher_path = output_dir / f"{voucher_file_path.stem}_AI补录版_{run_id}.xlsx"
    modified_ledger_path = output_dir / f"{ledger_file_path.stem}_AI补录版_{run_id}.xlsx"

    shutil.copy2(voucher_file_path, modified_voucher_path)
    shutil.copy2(ledger_file_path, modified_ledger_path)

    applied: list[dict[str, Any]] = []
    skipped: list[dict[str, Any]] = []
    valid_suggestions: list[dict[str, Any]] = []

    for suggestion in suggestions:
        if not _is_truthy(suggestion.get("apply_to_draft", True)):
            skipped.append({
                "suggestion": suggestion,
                "reason": "LLM建议不写入草稿"
            })
            continue

        ok, reason = _validate_suggestion(suggestion, valid_subject_codes)
        if not ok:
            skipped.append({
                "suggestion": suggestion,
                "reason": reason
            })
            continue

        valid_suggestions.append(suggestion)

    valid_suggestions.sort(key=lambda x: _normalize_date(x.get("transaction_date")))

    voucher_wb = load_workbook(modified_voucher_path)
    ledger_wb = load_workbook(modified_ledger_path)

    voucher_active_ws = voucher_wb.active
    ledger_active_ws = ledger_wb.active

    if voucher_active_ws is None or not isinstance(voucher_active_ws, Worksheet):
        raise ValueError("凭证草稿未找到可用工作表")

    if ledger_active_ws is None or not isinstance(ledger_active_ws, Worksheet):
        raise ValueError("台账草稿未找到可用工作表")

    voucher_ws = cast(Worksheet, voucher_active_ws)
    ledger_ws = cast(Worksheet, ledger_active_ws)

    voucher_header_row, voucher_header_map = _voucher_header_info(voucher_ws)
    ledger_header_row, ledger_header_map = _ledger_header_info(ledger_ws)

    for suggestion in valid_suggestions:
        try:
            _insert_voucher_pair(
                ws=voucher_ws,
                header_row=voucher_header_row,
                header_map=voucher_header_map,
                suggestion=suggestion,
                maker=maker,
                book_code=book_code,
                voucher_type=voucher_type,
            )

            applied.append({
                "flow_index": suggestion.get("flow_index", ""),
                "direction": suggestion.get("direction", ""),
                "transaction_date": _normalize_date(suggestion.get("transaction_date")),
                "counterparty": suggestion.get("counterparty", ""),
                "amount": _to_float(suggestion.get("amount")),
                "candidate_subject_code": _clean_code(suggestion.get("candidate_subject_code")),
                "candidate_subject_name": suggestion.get("candidate_subject_name", ""),
                "suggested_summary": _normalize_ai_summary(suggestion),
                "ledger_tag": suggestion.get("ledger_tag", ""),
                "confidence": suggestion.get("confidence", ""),
                "review_prompt": suggestion.get("review_prompt", ""),
            })

        except Exception as exc:
            skipped.append({
                "suggestion": suggestion,
                "reason": f"凭证写入失败：{str(exc)}"
            })

    _renumber_voucher_custom_no(voucher_ws, voucher_header_row, voucher_header_map)

    applied_flow_indices = {str(item.get("flow_index", "")) for item in applied}

    ledger_suggestions = [
        suggestion
        for suggestion in valid_suggestions
        if str(suggestion.get("flow_index", "")) in applied_flow_indices
    ]

    try:
        _insert_ledger_rows_before_total(
            ws=ledger_ws,
            header_row=ledger_header_row,
            header_map=ledger_header_map,
            suggestions=ledger_suggestions,
        )

        voucher_order = _get_voucher_summary_order(voucher_ws, voucher_header_row, voucher_header_map)

        _reorder_ledger_by_voucher_order(
            ws=ledger_ws,
            header_row=ledger_header_row,
            header_map=ledger_header_map,
            voucher_summary_order=voucher_order,
        )

        _recalc_ledger_total_and_income_composition(
            ws=ledger_ws,
            header_row=ledger_header_row,
            header_map=ledger_header_map,
        )

    except Exception as exc:
        skipped.append({
            "suggestion": {"flow_index": "ledger"},
            "reason": f"台账写入或重算失败：{str(exc)}"
        })

    voucher_wb.save(modified_voucher_path)
    ledger_wb.save(modified_ledger_path)

    return {
        "applied_count": len(applied),
        "skipped_count": len(skipped),
        "applied_suggestions": applied,
        "skipped_suggestions": skipped,
        "modified_voucher_file_name": modified_voucher_path.name,
        "modified_voucher_file_path": str(modified_voucher_path),
        "modified_ledger_file_name": modified_ledger_path.name,
        "modified_ledger_file_path": str(modified_ledger_path),
    }
