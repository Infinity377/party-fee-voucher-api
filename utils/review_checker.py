from __future__ import annotations

from pathlib import Path
from typing import Any, cast
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from utils.cleaners import normalize_cell
from utils.matcher import match_all_business
from utils.voucher_generator import (
    find_header_row as find_voucher_header_row,
    get_template_column_map as get_voucher_column_map,
    normalize_amount_for_excel,
)
from utils.ledger_generator import (
    find_ledger_header_row,
    get_ledger_column_map,
)


# =========================
# 基础工具
# =========================

def normalize_amount(value: Any) -> float:
    return normalize_amount_for_excel(value)


def amounts_equal(a: Any, b: Any, tolerance: float = 0.005) -> bool:
    return abs(normalize_amount(a) - normalize_amount(b)) <= tolerance


def normalize_date_text(value: Any) -> str:
    text = normalize_cell(value)
    if not text:
        return ""
    try:
        return pd.to_datetime(text).strftime("%Y-%m-%d")
    except Exception:
        return text[:10]


def normalize_compare_text(value: Any) -> str:
    return normalize_cell(value).replace(" ", "").replace("\n", "").replace("\r", "")


def normalize_code(value: Any) -> str:
    text = normalize_cell(value)
    if text.endswith(".0"):
        text = text[:-2]
    return text


def month_end_date(value: Any) -> str:
    text = normalize_date_text(value)
    if not text:
        return ""
    try:
        dt = pd.to_datetime(text)
        return (dt + pd.offsets.MonthEnd(0)).strftime("%Y-%m-%d")
    except Exception:
        return ""


def month_label(value: Any) -> str:
    text = normalize_date_text(value)
    if not text:
        return ""
    try:
        dt = pd.to_datetime(text)
        return f"{int(dt.month)}月"
    except Exception:
        return ""


def compact_text(value: Any) -> str:
    text = normalize_compare_text(value)
    text = text.replace("“", "").replace("”", "").replace('"', "")
    text = text.replace("《", "").replace("》", "")
    text = text.replace("（", "").replace("）", "").replace("(", "").replace(")", "")
    text = text.replace("，", "").replace(",", "").replace("。", "").replace("、", "")
    return text


def contains_compact(haystack: Any, needle: Any) -> bool:
    needle_text = compact_text(needle)
    haystack_text = compact_text(haystack)
    if not needle_text:
        return True
    return needle_text in haystack_text


def is_ai_text(value: Any) -> bool:
    text = normalize_compare_text(value)
    return "AI" in text.upper() or "待复核" in text or "补录" in text


def get_subject_lookup(subject_df: pd.DataFrame) -> dict[str, str]:
    lookup: dict[str, str] = {}
    if subject_df is None or subject_df.empty:
        return lookup

    for _, row in subject_df.iterrows():
        code = normalize_code(row.get("subject_code"))
        name = normalize_cell(row.get("subject_name"))
        if code:
            lookup[code] = name

    return lookup


def get_member_lookup(member_df: pd.DataFrame) -> dict[str, dict[str, str]]:
    lookup: dict[str, dict[str, str]] = {}
    if member_df is None or member_df.empty:
        return lookup

    for _, row in member_df.iterrows():
        name = normalize_cell(row.get("member_name"))
        if name:
            lookup[name] = {
                "member_status": normalize_cell(row.get("member_status")),
                "income_subject_code": normalize_code(row.get("income_subject_code")),
                "income_subject_name": normalize_cell(row.get("income_subject_name")),
            }

    return lookup


def get_oa_lookup(oa_df: pd.DataFrame) -> dict[str, dict[str, Any]]:
    lookup: dict[str, dict[str, Any]] = {}
    if oa_df is None or oa_df.empty:
        return lookup

    for _, row in oa_df.iterrows():
        flow_no = normalize_cell(row.get("oa_flow_no")).upper()
        if flow_no:
            lookup[flow_no] = {str(k): v for k, v in row.to_dict().items()}

    return lookup


# =========================
# Excel 读取
# =========================

def read_voucher_rows_from_excel(voucher_path: Path) -> list[dict[str, Any]]:
    wb = load_workbook(voucher_path, data_only=True)
    active_ws = wb.active
    if active_ws is None or not isinstance(active_ws, Worksheet):
        raise ValueError("凭证草稿未找到可用工作表")
    ws = cast(Worksheet, active_ws)

    header_row = find_voucher_header_row(ws)
    data_start_row = header_row + 1
    col_map = get_voucher_column_map(ws, header_row)

    fields = [
        "核算账簿", "凭证类别", "凭证号", "附单据数", "制单人", "制单日期",
        "审核人", "审核日期", "摘要", "表头自定义项2", "表头自定义项3",
        "科目编码", "币种", "原币借方金额", "本币借方金额",
        "原币贷方金额", "本币贷方金额", "票据号", "结算日期",
        "结算方式", "核销号", "核销业务日期",
    ]

    rows: list[dict[str, Any]] = []

    for row_idx in range(data_start_row, ws.max_row + 1):
        row_dict: dict[str, Any] = {"excel_row_no": row_idx}
        has_value = False

        for field in fields:
            col_idx = col_map.get(field)
            value = ""
            if col_idx is not None:
                value = ws.cell(row=row_idx, column=col_idx).value
            if normalize_cell(value):
                has_value = True
            row_dict[field] = value

        if has_value:
            rows.append(row_dict)

    return rows


def read_ledger_rows_from_excel(ledger_path: Path) -> list[dict[str, Any]]:
    wb = load_workbook(ledger_path, data_only=True)
    active_ws = wb.active
    if active_ws is None or not isinstance(active_ws, Worksheet):
        raise ValueError("台账草稿未找到可用工作表")
    ws = cast(Worksheet, active_ws)

    header_row = find_ledger_header_row(ws)
    data_start_row = header_row + 1
    col_map = get_ledger_column_map(ws, header_row)

    fields = [
        "年", "月", "日", "编号", "摘要", "支出", "收入", "余额", "标签",
        "科目编码", "科目名称", "流程编号", "对方单位", "收支方向",
    ]

    rows: list[dict[str, Any]] = []

    for row_idx in range(data_start_row, ws.max_row + 1):
        row_dict: dict[str, Any] = {"excel_row_no": row_idx}

        for field in fields:
            col_idx = col_map.get(field)
            value = ""
            if col_idx is not None:
                value = ws.cell(row=row_idx, column=col_idx).value
            row_dict[field] = value

        summary = normalize_compare_text(row_dict.get("摘要"))
        if summary in ["合计", "收入组成：", "收入组成"]:
            break

        has_value = any(
            normalize_cell(row_dict.get(field))
            for field in ["年", "月", "日", "摘要", "支出", "收入", "标签"]
        )
        if has_value:
            rows.append(row_dict)

    return rows


def read_ledger_income_composition(ledger_path: Path) -> dict[str, float]:
    wb = load_workbook(ledger_path, data_only=True)
    active_ws = wb.active
    if active_ws is None or not isinstance(active_ws, Worksheet):
        return {}
    ws = cast(Worksheet, active_ws)

    header_row = find_ledger_header_row(ws)
    col_map = get_ledger_column_map(ws, header_row)

    summary_col = col_map.get("摘要")
    income_col = col_map.get("收入")
    if summary_col is None or income_col is None:
        return {}

    found_title = False
    result: dict[str, float] = {}

    for row_idx in range(header_row + 1, ws.max_row + 1):
        label = normalize_cell(ws.cell(row=row_idx, column=summary_col).value)
        if label in {"收入组成：", "收入组成"}:
            found_title = True
            continue

        if not found_title:
            continue

        if not label:
            continue

        result[label] = normalize_amount(ws.cell(row=row_idx, column=income_col).value)

    return result


# =========================
# 业务单元构造
# =========================

def build_expected_units(
    matched_df: pd.DataFrame,
    original_business_exceptions: list[dict[str, Any]],
    oa_df: pd.DataFrame,
    member_df: pd.DataFrame,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """
    返回：
    1. 基础规则匹配成功的应复核业务单元
    2. 可接受的 AI 补录业务单元
    """
    oa_lookup = get_oa_lookup(oa_df)
    member_lookup = get_member_lookup(member_df)

    expected_units: list[dict[str, Any]] = []

    matched_only = matched_df[matched_df["match_status"] == "matched"].copy()
    matched_only = matched_only.reset_index(drop=True)

    for _, row in matched_only.iterrows():
        direction = normalize_cell(row.get("direction"))
        transaction_date = normalize_date_text(row.get("transaction_date"))
        voucher_date = normalize_date_text(row.get("voucher_date")) or month_end_date(transaction_date)
        oa_flow_no = normalize_cell(row.get("oa_flow_no")).upper()
        counterparty = normalize_cell(row.get("counterparty"))
        subject_code = normalize_code(row.get("subject_code"))

        application_title = ""
        if oa_flow_no and oa_flow_no in oa_lookup:
            application_title = normalize_cell(oa_lookup[oa_flow_no].get("application_title"))

        member_status = ""
        if counterparty in member_lookup:
            member_status = member_lookup[counterparty].get("member_status", "")

        expected_units.append({
            "source": "base_rule",
            "flow_index": row.get("flow_index"),
            "direction": direction,
            "transaction_date": transaction_date,
            "voucher_date": voucher_date,
            "amount": normalize_amount(row.get("amount")),
            "counterparty": counterparty,
            "oa_flow_no": oa_flow_no,
            "application_title": application_title,
            "subject_code": subject_code,
            "subject_name": normalize_cell(row.get("subject_name")),
            "summary": normalize_cell(row.get("summary")),
            "ledger_tag": normalize_cell(row.get("ledger_tag")),
            "business_type": normalize_cell(row.get("business_type")),
            "rule_id": normalize_cell(row.get("rule_id")),
            "member_status": member_status,
        })

    ai_allowed_units: list[dict[str, Any]] = []

    # 允许 AI 补录覆盖的原始异常业务：当前主要是易方达收入兜底。
    for item in original_business_exceptions:
        direction = normalize_cell(item.get("direction"))
        counterparty = normalize_cell(item.get("counterparty"))
        exception_text = normalize_cell(item.get("exception"))
        transaction_date = normalize_date_text(item.get("transaction_date"))
        amount = normalize_amount(item.get("amount"))

        if (
            direction == "收入"
            and "易方达" in counterparty
            and "收入流水无法匹配收入类规则" in exception_text
            and amount > 0
        ):
            ai_allowed_units.append({
                "source": "ai_allowed",
                "flow_index": item.get("flow_index"),
                "direction": "收入",
                "transaction_date": transaction_date,
                "voucher_date": month_end_date(transaction_date),
                "amount": amount,
                "counterparty": counterparty,
                "oa_flow_no": "",
                "application_title": "",
                "subject_code": "400101",
                "subject_name": "公司党员统一上缴党费",
                "summary": f"收到{counterparty}划来{month_label(transaction_date)}代收党费",
                "ledger_tag": "公司党员统一上缴",
                "business_type": "AI建议补录-易方达收入兜底",
                "rule_id": "AI-FALLBACK-YFD-INCOME",
                "member_status": "",
            })

    return expected_units, ai_allowed_units


def parse_voucher_business_units(actual_rows: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """
    将上传凭证两行合并为业务单元。
    """
    units: list[dict[str, Any]] = []
    exceptions: list[dict[str, Any]] = []

    if len(actual_rows) % 2 != 0:
        exceptions.append({
            "type": "VOUCHER_PAIR_INCOMPLETE",
            "message": f"上传凭证行数为 {len(actual_rows)}，不是偶数，可能存在借贷行缺失。",
            "suggestion": "请检查每笔业务是否均包含借方和贷方两行。"
        })

    pair_count = len(actual_rows) // 2

    for i in range(pair_count):
        row1 = actual_rows[i * 2]
        row2 = actual_rows[i * 2 + 1]

        rows = [row1, row2]
        debit_rows = [r for r in rows if normalize_amount(r.get("本币借方金额")) > 0 or normalize_amount(r.get("原币借方金额")) > 0]
        credit_rows = [r for r in rows if normalize_amount(r.get("本币贷方金额")) > 0 or normalize_amount(r.get("原币贷方金额")) > 0]

        debit_total = sum(normalize_amount(r.get("本币借方金额")) for r in rows)
        credit_total = sum(normalize_amount(r.get("本币贷方金额")) for r in rows)

        subject_codes = [normalize_code(r.get("科目编码")) for r in rows]
        debit_subjects = [normalize_code(r.get("科目编码")) for r in debit_rows]
        credit_subjects = [normalize_code(r.get("科目编码")) for r in credit_rows]

        direction = "未知"
        business_subject_code = ""

        if "1002" in debit_subjects and any(code != "1002" for code in credit_subjects):
            direction = "收入"
            business_subject_code = next((code for code in credit_subjects if code != "1002"), "")
        elif "1002" in credit_subjects and any(code != "1002" for code in debit_subjects):
            direction = "支出"
            business_subject_code = next((code for code in debit_subjects if code != "1002"), "")

        summary = normalize_cell(row1.get("摘要") or row2.get("摘要"))
        settlement_date = normalize_date_text(row1.get("结算日期") or row2.get("结算日期"))
        voucher_date = normalize_date_text(row1.get("制单日期") or row2.get("制单日期"))
        ai_mark = normalize_cell(row1.get("表头自定义项2") or row2.get("表头自定义项2"))

        units.append({
            "voucher_unit_index": i + 1,
            "excel_row_nos": [row1.get("excel_row_no"), row2.get("excel_row_no")],
            "direction": direction,
            "summary": summary,
            "settlement_date": settlement_date,
            "voucher_date": voucher_date,
            "amount": max(debit_total, credit_total),
            "debit_total": round(debit_total, 2),
            "credit_total": round(credit_total, 2),
            "subject_code": business_subject_code,
            "all_subject_codes": subject_codes,
            "ai_mark": ai_mark,
            "is_ai_supplement": is_ai_text(ai_mark),
            "row1": row1,
            "row2": row2,
        })

    return units, exceptions


def parse_ledger_units(actual_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    units: list[dict[str, Any]] = []

    for idx, row in enumerate(actual_rows, start=1):
        income = normalize_amount(row.get("收入"))
        expense = normalize_amount(row.get("支出"))

        direction = "未知"
        amount = 0.0

        if income > 0 and expense <= 0:
            direction = "收入"
            amount = income
        elif expense > 0 and income <= 0:
            direction = "支出"
            amount = expense
        elif income > 0 and expense > 0:
            direction = "双向金额异常"
            amount = max(income, expense)

        year = normalize_cell(row.get("年"))
        month = normalize_cell(row.get("月")).zfill(2) if normalize_cell(row.get("月")) else ""
        day = normalize_cell(row.get("日")).zfill(2) if normalize_cell(row.get("日")) else ""
        ledger_date = f"{year}-{month}-{day}" if year and month and day else ""

        units.append({
            "ledger_unit_index": idx,
            "excel_row_no": row.get("excel_row_no"),
            "direction": direction,
            "ledger_date": ledger_date,
            "summary": normalize_cell(row.get("摘要")),
            "amount": amount,
            "income": income,
            "expense": expense,
            "tag": normalize_cell(row.get("标签")),
            "raw": row,
        })

    return units


# =========================
# 关键词复核
# =========================

def extract_org_terms(text: str) -> list[str]:
    text = normalize_cell(text)
    terms: list[str] = []

    # 党组织名称
    for m in re.finditer(r"([\u4e00-\u9fa5A-Za-z0-9与和、（）()\"“”《》]+?(?:党支部|党总支|党委|党组织))", text):
        term = m.group(1)
        term = re.sub(r"^(关于|报销|申请|支付)", "", term)
        term = term.strip("，,。:：；; ")
        if len(term) >= 3:
            terms.append(term)

    # 引号或书名号中的项目名称
    for m in re.finditer(r"[\"“《](.*?)[\"”》]", text):
        term = m.group(1).strip()
        if len(term) >= 2:
            terms.append(term)

    # 赴某地开展
    for m in re.finditer(r"赴([\u4e00-\u9fa5A-Za-z0-9]+?)开展", text):
        term = m.group(1).strip()
        if len(term) >= 2:
            terms.append(term)

    # 常见关键地点/对象
    for keyword in ["延安", "南昌", "学校", "A单位", "求是", "人民日报"]:
        if keyword in text:
            terms.append(keyword)

    # 会议名称
    meeting_match = re.search(r"([\u4e00-\u9fa5A-Za-z0-9（）()]+会议)", text)
    if meeting_match:
        term = meeting_match.group(1).strip()
        if len(term) >= 3:
            terms.append(term)

    # 去重，保序
    result: list[str] = []
    for term in terms:
        if term and term not in result:
            result.append(term)
    return result


def expected_critical_terms(unit: dict[str, Any]) -> list[str]:
    terms: list[str] = []

    direction = normalize_cell(unit.get("direction"))
    counterparty = normalize_cell(unit.get("counterparty"))
    summary = normalize_cell(unit.get("summary"))
    oa_flow_no = normalize_cell(unit.get("oa_flow_no")).upper()
    application_title = normalize_cell(unit.get("application_title"))
    business_type = normalize_cell(unit.get("business_type"))

    if direction == "收入":
        if counterparty:
            terms.append(counterparty)
        if "利息" in business_type or "利息" in summary:
            terms.append("利息")
            period_match = re.search(r"(\d{4}-\d{1,2}-\d{1,2}到\d{4}-\d{1,2}-\d{1,2})", summary)
            if period_match:
                terms.append(period_match.group(1))
    elif direction == "支出":
        if oa_flow_no:
            terms.append(oa_flow_no)
        terms.extend(extract_org_terms(application_title))
        # 对清洗后摘要中的关键组织名称也补充检查
        terms.extend(extract_org_terms(summary))

    # 去掉过短或容易误判的词
    filtered: list[str] = []
    for term in terms:
        term = normalize_cell(term)
        if len(compact_text(term)) < 2:
            continue
        if term not in filtered:
            filtered.append(term)

    return filtered


def find_missing_critical_terms(expected_unit: dict[str, Any], actual_summary: str) -> list[str]:
    missing: list[str] = []
    for term in expected_critical_terms(expected_unit):
        if not contains_compact(actual_summary, term):
            missing.append(term)
    return missing


# =========================
# 匹配实际凭证/台账
# =========================

def score_voucher_unit_match(expected: dict[str, Any], actual: dict[str, Any]) -> int:
    score = 0

    if normalize_cell(expected.get("direction")) == normalize_cell(actual.get("direction")):
        score += 4

    if normalize_date_text(expected.get("transaction_date")) == normalize_date_text(actual.get("settlement_date")):
        score += 4

    if amounts_equal(expected.get("amount"), actual.get("amount")):
        score += 3

    summary = normalize_cell(actual.get("summary"))
    flow_no = normalize_cell(expected.get("oa_flow_no")).upper()
    counterparty = normalize_cell(expected.get("counterparty"))

    if flow_no and contains_compact(summary, flow_no):
        score += 5

    if counterparty and contains_compact(summary, counterparty):
        score += 5

    # 即使摘要被改错，也允许日期+金额+方向匹配，便于继续报摘要错误。
    return score


def find_best_voucher_unit(
    expected: dict[str, Any],
    actual_units: list[dict[str, Any]],
    used_indexes: set[int],
    min_score: int = 8,
) -> dict[str, Any] | None:
    best_unit = None
    best_score = -1

    for idx, actual in enumerate(actual_units):
        if idx in used_indexes:
            continue

        score = score_voucher_unit_match(expected, actual)
        if score > best_score:
            best_score = score
            best_unit = actual

    if best_unit is not None and best_score >= min_score:
        used_indexes.add(actual_units.index(best_unit))
        return best_unit

    return None


def score_ledger_unit_match(expected: dict[str, Any], actual: dict[str, Any]) -> int:
    score = 0

    if normalize_cell(expected.get("direction")) == normalize_cell(actual.get("direction")):
        score += 4

    if normalize_date_text(expected.get("voucher_date")) == normalize_date_text(actual.get("ledger_date")):
        score += 4

    if amounts_equal(expected.get("amount"), actual.get("amount")):
        score += 3

    summary = normalize_cell(actual.get("summary"))
    counterparty = normalize_cell(expected.get("counterparty"))
    oa_flow_no = normalize_cell(expected.get("oa_flow_no")).upper()

    if counterparty and contains_compact(summary, counterparty):
        score += 5

    if oa_flow_no and contains_compact(summary, oa_flow_no):
        score += 5

    return score


def find_best_ledger_unit(
    expected: dict[str, Any],
    actual_units: list[dict[str, Any]],
    used_indexes: set[int],
    min_score: int = 8,
) -> dict[str, Any] | None:
    best_unit = None
    best_score = -1

    for idx, actual in enumerate(actual_units):
        if idx in used_indexes:
            continue

        score = score_ledger_unit_match(expected, actual)
        if score > best_score:
            best_score = score
            best_unit = actual

    if best_unit is not None and best_score >= min_score:
        used_indexes.add(actual_units.index(best_unit))
        return best_unit

    return None


# =========================
# 具体复核
# =========================

def check_voucher_balance(actual_rows: list[dict[str, Any]]) -> dict[str, Any]:
    debit_total = 0.0
    credit_total = 0.0

    for row in actual_rows:
        debit_total += normalize_amount(row.get("本币借方金额"))
        credit_total += normalize_amount(row.get("本币贷方金额"))

    balance_check = abs(debit_total - credit_total) <= 0.005

    return {
        "debit_total": round(debit_total, 2),
        "credit_total": round(credit_total, 2),
        "balance_check": balance_check,
    }


def review_voucher_units(
    expected_units: list[dict[str, Any]],
    ai_allowed_units: list[dict[str, Any]],
    actual_units: list[dict[str, Any]],
    subject_lookup: dict[str, str],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    exceptions: list[dict[str, Any]] = []
    ai_supplement_items: list[dict[str, Any]] = []
    used_indexes: set[int] = set()

    all_expected = [*expected_units, *ai_allowed_units]

    for expected in all_expected:
        actual = find_best_voucher_unit(expected, actual_units, used_indexes)
        is_ai_expected = expected.get("source") == "ai_allowed"

        if actual is None:
            # 原始异常不强制要求 AI 补录；只有基础 matched 业务缺凭证才算异常。
            if not is_ai_expected:
                exceptions.append({
                    "type": "VOUCHER_BUSINESS_MISSING",
                    "message": (
                        f"未在上传凭证中找到流水序号{expected.get('flow_index')}对应业务："
                        f"日期{expected.get('transaction_date')}，方向{expected.get('direction')}，"
                        f"金额{expected.get('amount')}，对方单位/户名：{expected.get('counterparty')}。"
                    ),
                    "suggestion": "请检查凭证是否漏填或被误删。"
                })
            continue

        # 基础字段
        if normalize_cell(actual.get("direction")) != normalize_cell(expected.get("direction")):
            exceptions.append({
                "type": "VOUCHER_DIRECTION_MISMATCH",
                "message": (
                    f"流水序号{expected.get('flow_index')}凭证借贷方向不一致："
                    f"应为{expected.get('direction')}，实际识别为{actual.get('direction')}。"
                ),
                "suggestion": "收入应借记银行存款、贷记收入科目；支出应借记支出科目、贷记银行存款。"
            })

        if not amounts_equal(expected.get("amount"), actual.get("amount")):
            exceptions.append({
                "type": "VOUCHER_AMOUNT_MISMATCH",
                "message": (
                    f"流水序号{expected.get('flow_index')}凭证金额不一致："
                    f"流水金额应为 {expected.get('amount')}，上传凭证金额为 {actual.get('amount')}。"
                ),
                "suggestion": "请检查凭证借贷金额是否与银行流水金额一致。"
            })

        if normalize_date_text(expected.get("transaction_date")) != normalize_date_text(actual.get("settlement_date")):
            exceptions.append({
                "type": "VOUCHER_SETTLEMENT_DATE_MISMATCH",
                "message": (
                    f"流水序号{expected.get('flow_index')}凭证结算日期不一致："
                    f"应为 {expected.get('transaction_date')}，实际为 {actual.get('settlement_date')}。"
                ),
                "suggestion": "结算日期、核销业务日期应对应银行流水交易时间。"
            })

        expected_subject = normalize_code(expected.get("subject_code"))
        actual_subject = normalize_code(actual.get("subject_code"))

        if actual_subject and actual_subject not in subject_lookup:
            exceptions.append({
                "type": "VOUCHER_SUBJECT_NOT_IN_TABLE",
                "message": f"凭证科目编码 {actual_subject} 不存在于会计科目表。",
                "suggestion": "请使用会计科目表中的末级科目编码。"
            })

        if expected_subject and actual_subject != expected_subject:
            exceptions.append({
                "type": "VOUCHER_SUBJECT_RULE_MISMATCH",
                "message": (
                    f"流水序号{expected.get('flow_index')}凭证科目编码不符合规则："
                    f"根据业务映射规则/党员状态表应为 {expected_subject} {expected.get('subject_name')}，"
                    f"实际为 {actual_subject} {subject_lookup.get(actual_subject, '')}。"
                ),
                "suggestion": "请根据党费业务映射规则表、党员离退休情况表和会计科目表修正科目编码。"
            })

        missing_terms = find_missing_critical_terms(expected, normalize_cell(actual.get("summary")))
        if missing_terms:
            exceptions.append({
                "type": "VOUCHER_SUMMARY_KEYWORD_MISMATCH",
                "message": (
                    f"流水序号{expected.get('flow_index')}凭证摘要缺少关键业务信息："
                    f"{'、'.join(missing_terms)}。当前摘要为“{actual.get('summary')}”。"
                ),
                "suggestion": "请确保凭证摘要中的姓名、公司、党支部/党总支、流程编号、活动地点或利息期间与流水/OA流程一致。"
            })

        if is_ai_expected:
            if not actual.get("is_ai_supplement"):
                exceptions.append({
                    "type": "AI_SUPPLEMENT_MARK_MISSING",
                    "message": (
                        f"流水序号{expected.get('flow_index')}属于AI建议补录业务，"
                        f"但凭证未标注“AI建议补录-待复核”。"
                    ),
                    "suggestion": "AI补录分录必须明确标注待人工复核，避免被误认为正式自动入账。"
                })
            else:
                ai_supplement_items.append({
                    "flow_index": expected.get("flow_index"),
                    "transaction_date": expected.get("transaction_date"),
                    "direction": expected.get("direction"),
                    "counterparty": expected.get("counterparty"),
                    "amount": expected.get("amount"),
                    "candidate_subject_code": expected.get("subject_code"),
                    "candidate_subject_name": expected.get("subject_name"),
                    "summary": actual.get("summary"),
                    "review_prompt": "该业务为AI建议补录，不计入复核异常，但需财务人员核实该款项性质后方可正式入账。"
                })

    return exceptions, ai_supplement_items


def review_ledger_units(
    expected_units: list[dict[str, Any]],
    ai_allowed_units: list[dict[str, Any]],
    actual_units: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    exceptions: list[dict[str, Any]] = []
    ai_supplement_items: list[dict[str, Any]] = []
    used_indexes: set[int] = set()

    for actual in actual_units:
        if actual["income"] > 0 and actual["expense"] > 0:
            exceptions.append({
                "type": "LEDGER_AMOUNT_DIRECTION_CONFLICT",
                "message": (
                    f"台账第 {actual.get('excel_row_no')} 行同时存在收入和支出金额："
                    f"收入 {actual['income']}，支出 {actual['expense']}。"
                ),
                "suggestion": "收入业务只能填写收入列，支出业务只能填写支出列。"
            })

    all_expected = [*expected_units, *ai_allowed_units]

    for expected in all_expected:
        actual = find_best_ledger_unit(expected, actual_units, used_indexes)
        is_ai_expected = expected.get("source") == "ai_allowed"

        if actual is None:
            if not is_ai_expected:
                exceptions.append({
                    "type": "LEDGER_BUSINESS_MISSING",
                    "message": (
                        f"未在上传台账中找到流水序号{expected.get('flow_index')}对应业务："
                        f"日期{expected.get('voucher_date')}，方向{expected.get('direction')}，"
                        f"金额{expected.get('amount')}，对方单位/户名：{expected.get('counterparty')}。"
                    ),
                    "suggestion": "请检查台账明细是否漏填或被误删。"
                })
            continue

        if normalize_cell(actual.get("direction")) != normalize_cell(expected.get("direction")):
            exceptions.append({
                "type": "LEDGER_DIRECTION_MISMATCH",
                "message": (
                    f"流水序号{expected.get('flow_index')}台账收支方向不一致："
                    f"应为{expected.get('direction')}，实际为{actual.get('direction')}。"
                ),
                "suggestion": "请检查该业务应填收入列还是支出列。"
            })

        if not amounts_equal(expected.get("amount"), actual.get("amount")):
            exceptions.append({
                "type": "LEDGER_AMOUNT_MISMATCH",
                "message": (
                    f"流水序号{expected.get('flow_index')}台账金额不一致："
                    f"应为 {expected.get('amount')}，实际为 {actual.get('amount')}。"
                ),
                "suggestion": "请检查台账收入/支出金额是否与银行流水和凭证一致。"
            })

        if normalize_date_text(expected.get("voucher_date")) != normalize_date_text(actual.get("ledger_date")):
            exceptions.append({
                "type": "LEDGER_DATE_MISMATCH",
                "message": (
                    f"流水序号{expected.get('flow_index')}台账日期不一致："
                    f"应为 {expected.get('voucher_date')}，实际为 {actual.get('ledger_date')}。"
                ),
                "suggestion": "台账年月日应对应凭证制单日期，即交易月份月末。"
            })

        expected_tag = normalize_cell(expected.get("ledger_tag"))
        actual_tag = normalize_cell(actual.get("tag"))

        if expected_tag and actual_tag != expected_tag:
            exceptions.append({
                "type": "LEDGER_TAG_RULE_MISMATCH",
                "message": (
                    f"流水序号{expected.get('flow_index')}台账标签不符合规则："
                    f"根据业务映射规则/党员状态表应为“{expected_tag}”，实际为“{actual_tag}”。"
                ),
                "suggestion": "请根据党费业务映射规则表修正台账标签。利息收入应为“党费账户利息收入”，公司代收应为“公司党员统一上缴”，个人缴费应为“其他党员自行上缴”。"
            })

        missing_terms = find_missing_critical_terms(expected, normalize_cell(actual.get("summary")))
        if missing_terms:
            exceptions.append({
                "type": "LEDGER_SUMMARY_KEYWORD_MISMATCH",
                "message": (
                    f"流水序号{expected.get('flow_index')}台账摘要缺少关键业务信息："
                    f"{'、'.join(missing_terms)}。当前摘要为“{actual.get('summary')}”。"
                ),
                "suggestion": "请确保台账摘要中的姓名、公司、党支部/党总支、流程编号、活动地点或利息期间与流水/OA流程一致。"
            })

        if is_ai_expected and actual is not None:
            ai_supplement_items.append({
                "flow_index": expected.get("flow_index"),
                "transaction_date": expected.get("transaction_date"),
                "direction": expected.get("direction"),
                "counterparty": expected.get("counterparty"),
                "amount": expected.get("amount"),
                "candidate_subject_code": expected.get("subject_code"),
                "candidate_subject_name": expected.get("subject_name"),
                "summary": actual.get("summary"),
                "review_prompt": "该台账行为AI建议补录，不计入复核异常，但需财务人员核实后方可正式使用。"
            })

    return exceptions, ai_supplement_items


def review_income_composition(
    actual_ledger_units: list[dict[str, Any]],
    actual_composition: dict[str, float],
) -> list[dict[str, Any]]:
    exceptions: list[dict[str, Any]] = []

    if not actual_composition:
        return exceptions

    tags = ["公司党员统一上缴", "其他党员自行上缴", "广发下拨", "党费账户利息收入"]

    detail_by_tag = {tag: 0.0 for tag in tags}
    for unit in actual_ledger_units:
        if unit.get("direction") == "收入":
            tag = normalize_cell(unit.get("tag"))
            if tag in detail_by_tag:
                detail_by_tag[tag] += normalize_amount(unit.get("amount"))

    for tag in tags:
        detail_value = round(detail_by_tag.get(tag, 0.0), 2)
        composition_value = round(actual_composition.get(tag, 0.0), 2)
        if not amounts_equal(detail_value, composition_value):
            exceptions.append({
                "type": "LEDGER_INCOME_COMPOSITION_MISMATCH",
                "message": (
                    f"台账收入组成【{tag}】不一致：明细汇总为 {detail_value}，"
                    f"收入组成区域为 {composition_value}。"
                ),
                "suggestion": "请重新计算收入组成，并确认各收入明细标签是否正确。"
            })

    expected_total = round(sum(detail_by_tag.values()), 2)
    actual_total = round(actual_composition.get("合计：", actual_composition.get("合计", 0.0)), 2)

    if not amounts_equal(expected_total, actual_total):
        exceptions.append({
            "type": "LEDGER_INCOME_COMPOSITION_TOTAL_MISMATCH",
            "message": f"台账收入组成合计不一致：明细汇总为 {expected_total}，收入组成合计为 {actual_total}。",
            "suggestion": "请检查收入组成区域合计是否等于各收入标签汇总。"
        })

    return exceptions


# =========================
# 报告文本
# =========================

def build_exception_summary(review_exceptions: list[dict[str, Any]], max_items: int = 5) -> str:
    if not review_exceptions:
        return ""

    lines = []
    for idx, item in enumerate(review_exceptions[:max_items], start=1):
        message = normalize_cell(item.get("message"))
        suggestion = normalize_cell(item.get("suggestion"))
        if suggestion:
            lines.append(f"{idx}）{message}建议：{suggestion}")
        else:
            lines.append(f"{idx}）{message}")

    remaining_count = len(review_exceptions) - max_items
    if remaining_count > 0:
        lines.append(f"其余 {remaining_count} 项异常详见“复核异常清单”工作表。")

    return "主要异常包括：" + "；".join(lines)


def build_ai_supplement_summary(ai_supplement_items: list[dict[str, Any]]) -> str:
    if not ai_supplement_items:
        return "未检测到AI补录业务。"

    return (
        f"检测到 {len(ai_supplement_items)} 条AI补录提示记录。"
        f"该类业务不计入复核异常，但必须由财务人员结合银行回单、OA流程、党费收缴明细或业务部门说明人工确认。"
    )


def build_review_report(
    review_status: str,
    matched_count: int,
    original_exception_count: int,
    voucher_balance: dict[str, Any],
    review_exceptions: list[dict[str, Any]],
    ai_supplement_items: list[dict[str, Any]],
) -> str:
    ai_text = build_ai_supplement_summary(ai_supplement_items)

    if review_status == "passed":
        return (
            f"系统复核无异常。系统已根据银行流水、OA流程、党费业务映射规则表、党员离退休情况表和会计科目表，"
            f"对上传凭证草稿及台账草稿进行业务实质复核。"
            f"本次基础规则成功匹配业务 {matched_count} 笔，原始异常业务 {original_exception_count} 笔；"
            f"凭证借方合计 {voucher_balance['debit_total']}，贷方合计 {voucher_balance['credit_total']}，借贷平衡校验通过。"
            f"{ai_text}仍建议由财务人员进行最终人工确认。"
        )

    exception_summary = build_exception_summary(review_exceptions, max_items=5)

    return (
        f"系统复核发现异常。系统已根据银行流水、OA流程、党费业务映射规则表、党员离退休情况表和会计科目表，"
        f"对上传凭证草稿及台账草稿进行业务实质复核，发现 {len(review_exceptions)} 项异常。"
        f"本次基础规则成功匹配业务 {matched_count} 笔，原始异常业务 {original_exception_count} 笔；"
        f"凭证借方合计 {voucher_balance['debit_total']}，贷方合计 {voucher_balance['credit_total']}，"
        f"借贷平衡校验结果为{'通过' if voucher_balance['balance_check'] else '不通过'}。"
        f"{ai_text}{exception_summary}"
        f"详细信息请查看复核报告 Excel 中的“复核异常清单”工作表，并根据异常清单逐项核对。"
    )


# =========================
# 主入口
# =========================

def perform_review(
    bank_df: pd.DataFrame,
    oa_df: pd.DataFrame,
    subject_df: pd.DataFrame,
    member_df: pd.DataFrame,
    rule_df: pd.DataFrame,
    voucher_path: Path,
    ledger_path: Path,
    maker: str = "",
    book_code: str = "",
    voucher_type: str = "",
) -> dict[str, Any]:
    matched_df, original_business_exceptions = match_all_business(
        bank_df=bank_df,
        oa_df=oa_df,
        subject_df=subject_df,
        member_df=member_df,
        rule_df=rule_df
    )

    matched_count = int((matched_df["match_status"] == "matched").sum())
    original_exception_count = int(len(original_business_exceptions))

    expected_units, ai_allowed_units = build_expected_units(
        matched_df=matched_df,
        original_business_exceptions=original_business_exceptions,
        oa_df=oa_df,
        member_df=member_df,
    )

    actual_voucher_rows = read_voucher_rows_from_excel(voucher_path)
    actual_ledger_rows = read_ledger_rows_from_excel(ledger_path)

    actual_voucher_units, voucher_parse_exceptions = parse_voucher_business_units(actual_voucher_rows)
    actual_ledger_units = parse_ledger_units(actual_ledger_rows)
    actual_income_composition = read_ledger_income_composition(ledger_path)

    subject_lookup = get_subject_lookup(subject_df)

    voucher_balance = check_voucher_balance(actual_voucher_rows)

    review_exceptions: list[dict[str, Any]] = []

    if not voucher_balance["balance_check"]:
        review_exceptions.append({
            "type": "VOUCHER_NOT_BALANCED",
            "message": f"上传凭证借贷不平：借方合计 {voucher_balance['debit_total']}，贷方合计 {voucher_balance['credit_total']}。",
            "suggestion": "请检查凭证金额是否漏填、错填，或是否误删凭证明细行。"
        })

    review_exceptions.extend(voucher_parse_exceptions)

    voucher_exceptions, voucher_ai_items = review_voucher_units(
        expected_units=expected_units,
        ai_allowed_units=ai_allowed_units,
        actual_units=actual_voucher_units,
        subject_lookup=subject_lookup,
    )
    review_exceptions.extend(voucher_exceptions)

    ledger_exceptions, ledger_ai_items = review_ledger_units(
        expected_units=expected_units,
        ai_allowed_units=ai_allowed_units,
        actual_units=actual_ledger_units,
    )
    review_exceptions.extend(ledger_exceptions)

    review_exceptions.extend(
        review_income_composition(
            actual_ledger_units=actual_ledger_units,
            actual_composition=actual_income_composition,
        )
    )

    ai_supplement_items = voucher_ai_items or ledger_ai_items

    # 去重：同一个 flow_index 的 AI 提示只保留一条
    deduped_ai_items: list[dict[str, Any]] = []
    seen_ai_keys: set[str] = set()
    for item in ai_supplement_items:
        key = str(item.get("flow_index"))
        if key in seen_ai_keys:
            continue
        seen_ai_keys.add(key)
        deduped_ai_items.append(item)

    review_status = "passed" if not review_exceptions else "failed"

    review_report = build_review_report(
        review_status=review_status,
        matched_count=matched_count,
        original_exception_count=original_exception_count,
        voucher_balance=voucher_balance,
        review_exceptions=review_exceptions,
        ai_supplement_items=deduped_ai_items,
    )

    return {
        "review_status": review_status,
        "review_passed": review_status == "passed",
        "review_report": review_report,
        "review_exception_count": len(review_exceptions),
        "review_exceptions": review_exceptions,
        "original_business_exception_count": original_exception_count,
        "original_business_exceptions": original_business_exceptions,
        "matched_count": matched_count,
        "expected_voucher_row_count": len(expected_units) * 2,
        "actual_voucher_row_count": len(actual_voucher_rows),
        "expected_ledger_row_count": len(expected_units),
        "actual_ledger_row_count": len(actual_ledger_rows),
        "ai_supplement_count": len(deduped_ai_items),
        "ai_supplement_items": deduped_ai_items,
        "voucher_balance": voucher_balance,
    }
