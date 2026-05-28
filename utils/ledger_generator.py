from __future__ import annotations

from pathlib import Path
from typing import Any, cast
import shutil
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import MergedCell

from utils.cleaners import normalize_cell


FIXED_INCOME_TAG_ORDER = [
    "公司党员统一上缴",
    "其他党员自行上缴",
    "广发下拨",
    "党费账户利息收入",
]


def normalize_amount_for_excel(value: Any) -> float:
    """
    将金额转为 Excel 可写入的 float。
    """
    if value is None or pd.isna(value):
        return 0.0

    if isinstance(value, (int, float)):
        return round(float(value), 2)

    text = str(value).strip().replace(",", "").replace("，", "")
    if text == "":
        return 0.0

    try:
        return round(float(text), 2)
    except ValueError:
        return 0.0


def normalize_header(value: Any) -> str:
    """
    清洗模板字段名。
    """
    text = normalize_cell(value)
    text = text.replace("*", "")
    text = text.replace("＊", "")
    text = text.replace("（元）", "")
    text = text.replace("(元)", "")
    text = text.replace(" ", "")
    return text.strip()


def split_year_month_day(date_text: Any) -> tuple[Any, Any, Any]:
    """
    将 YYYY-MM-DD 拆成 年、月、日。
    """
    text = normalize_cell(date_text)

    try:
        dt = pd.to_datetime(text)
        return int(dt.year), f"{int(dt.month):02d}", f"{int(dt.day):02d}"
    except Exception:
        parts = text.split("-")
        if len(parts) >= 3:
            return parts[0], parts[1], parts[2]
        return "", "", ""


def safe_set_cell_value(ws: Worksheet, row: int, column: int, value: Any) -> None:
    """
    安全写入单元格。
    如果目标单元格属于合并单元格，则写入该合并区域左上角单元格。
    """
    cell_obj = ws.cell(row=row, column=column)

    if not isinstance(cell_obj, MergedCell):
        cell_obj.value = value
        return

    for merged_range in ws.merged_cells.ranges:
        if cell_obj.coordinate in merged_range:
            top_left_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
            top_left_cell.value = value
            return


def unmerge_rows_from(ws: Worksheet, start_row: int) -> None:
    """
    取消 start_row 及其以下区域的合并单元格。
    """
    ranges_to_unmerge = []

    for merged_range in list(ws.merged_cells.ranges):
        if merged_range.max_row >= start_row:
            ranges_to_unmerge.append(str(merged_range))

    for range_text in ranges_to_unmerge:
        ws.unmerge_cells(range_text)


def find_ledger_header_row(ws: Worksheet) -> int:
    """
    自动寻找台账表头行。
    """
    max_scan_row = min(ws.max_row, 20)

    required_keywords = ["年", "月", "日", "摘要"]
    optional_keywords = ["收入", "支出", "余额", "标签", "编号"]

    best_row = 1
    best_score = 0

    for row_idx in range(1, max_scan_row + 1):
        row_values = [
            str(ws.cell(row=row_idx, column=col_idx).value or "")
            for col_idx in range(1, ws.max_column + 1)
        ]
        row_text = "|".join(row_values)

        score = 0

        for keyword in required_keywords:
            if keyword in row_text:
                score += 2

        for keyword in optional_keywords:
            if keyword in row_text:
                score += 1

        if score > best_score:
            best_score = score
            best_row = row_idx

    return best_row


def get_ledger_column_map(ws: Worksheet, header_row: int) -> dict[str, int]:
    """
    获取台账模板字段到列号的映射。
    兼容样例台账：
    A 年、B 月、C 日、D 编号、E 摘要、F 支出、G 收入、H 余额、I 标签。
    """
    raw_map: dict[str, int] = {}

    for col_idx in range(1, ws.max_column + 1):
        value = ws.cell(row=header_row, column=col_idx).value
        header = normalize_header(value)
        if header:
            raw_map[header] = col_idx

    field_map: dict[str, int] = {}

    for header, col_idx in raw_map.items():
        if header == "年":
            field_map["年"] = col_idx
        elif header == "月":
            field_map["月"] = col_idx
        elif header == "日":
            field_map["日"] = col_idx
        elif "编号" in header:
            field_map["编号"] = col_idx
        elif "摘要" in header or "事由" in header:
            field_map["摘要"] = col_idx
        elif "支出" in header or "贷方" in header:
            field_map["支出"] = col_idx
        elif "收入" in header or "借方" in header:
            field_map["收入"] = col_idx
        elif "余额" in header:
            field_map["余额"] = col_idx
        elif "标签" in header or "类别" in header:
            field_map["标签"] = col_idx
        elif "科目编码" in header:
            field_map["科目编码"] = col_idx
        elif "科目名称" in header:
            field_map["科目名称"] = col_idx
        elif "流程" in header or "OA" in header or "单号" in header:
            field_map["流程编号"] = col_idx
        elif "对方" in header or "单位" in header:
            field_map["对方单位"] = col_idx
        elif "方向" in header or "收支" in header:
            field_map["收支方向"] = col_idx

    return field_map


def clear_old_data(ws: Worksheet, start_row: int) -> None:
    """
    清空模板旧数据，保留表头。
    """
    unmerge_rows_from(ws, start_row)

    if ws.max_row >= start_row:
        ws.delete_rows(start_row, ws.max_row - start_row + 1)


def build_ledger_rows(matched_df: pd.DataFrame) -> list[dict[str, Any]]:
    """
    根据匹配后的业务明细生成台账行。
    只取 match_status = matched 的业务。
    异常业务不进入正式台账草稿。

    注意：
    H列“余额”不自动计算，保持为空。
    """
    rows: list[dict[str, Any]] = []

    matched_only = matched_df[matched_df["match_status"] == "matched"].copy()
    matched_only = matched_only.reset_index(drop=True)

    if "voucher_date" in matched_only.columns and "flow_index" in matched_only.columns:
        matched_only = matched_only.sort_values(by=["voucher_date", "flow_index"]).reset_index(drop=True)

    for _, row in matched_only.iterrows():
        direction = normalize_cell(row.get("direction"))
        amount = normalize_amount_for_excel(row.get("amount"))
        voucher_date = normalize_cell(row.get("voucher_date"))
        year, month, day = split_year_month_day(voucher_date)

        income_amount: float | str = ""
        expense_amount: float | str = ""

        if direction == "收入":
            income_amount = amount
        elif direction == "支出":
            expense_amount = amount
        else:
            continue

        ledger_row = {
            "年": year,
            "月": month,
            "日": day,
            "日期": voucher_date,
            "编号": "",
            "摘要": normalize_cell(row.get("summary")),
            "收入": income_amount,
            "支出": expense_amount,
            "余额": "",
            "标签": normalize_cell(row.get("ledger_tag")),
            "科目编码": normalize_cell(row.get("subject_code")),
            "科目名称": normalize_cell(row.get("subject_name")),
            "流程编号": normalize_cell(row.get("oa_flow_no")),
            "对方单位": normalize_cell(row.get("counterparty")),
            "收支方向": direction,
        }

        rows.append(ledger_row)

    return rows


def build_income_composition_rows(ledger_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """
    以“标签”为基准聚合收入金额，生成台账底部“收入组成”。
    默认固定展示：
    - 公司党员统一上缴
    - 其他党员自行上缴
    - 广发下拨
    - 党费账户利息收入
    - 合计
    """
    income_by_tag: dict[str, float] = {}

    for row in ledger_rows:
        tag = normalize_cell(row.get("标签"))
        income_amount = normalize_amount_for_excel(row.get("收入"))

        if income_amount > 0:
            income_by_tag[tag] = income_by_tag.get(tag, 0.0) + income_amount

    result: list[dict[str, Any]] = []

    for tag in FIXED_INCOME_TAG_ORDER:
        result.append(
            {
                "项目": tag,
                "收入": round(income_by_tag.get(tag, 0.0), 2),
            }
        )

    total_income = round(sum(float(item["收入"]) for item in result), 2)
    result.append(
        {
            "项目": "合计：",
            "收入": total_income,
        }
    )

    return result


def write_ledger_rows_to_template(
    ledger_rows: list[dict[str, Any]],
    template_path: Path,
    output_path: Path
) -> Path:
    """
    将台账行写入台账模板，并在明细下方追加合计与收入组成。
    """
    if not template_path.exists():
        raise FileNotFoundError(f"台账模板不存在：{template_path}")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copyfile(template_path, output_path)

    wb = load_workbook(output_path)
    ws = cast(Worksheet, wb.active)

    header_row = find_ledger_header_row(ws)
    data_start_row = header_row + 1
    col_map = get_ledger_column_map(ws, header_row)

    clear_old_data(ws, data_start_row)

    for row_offset, ledger_row in enumerate(ledger_rows):
        excel_row = data_start_row + row_offset

        # 先显式清空收入/支出/余额列，避免模板或复制行残留旧金额。
        for clear_field in ["收入", "支出", "余额"]:
            clear_col = col_map.get(clear_field)
            if clear_col is not None:
                safe_set_cell_value(ws, excel_row, clear_col, "")

        # 再写入当前业务行。build_ledger_rows 已保证收入行只填收入、支出行只填支出。
        for field_name, value in ledger_row.items():
            col_idx = col_map.get(field_name)

            if col_idx is None:
                continue

            safe_set_cell_value(ws, excel_row, col_idx, value)

    income_total = round(
        sum(normalize_amount_for_excel(row.get("收入")) for row in ledger_rows),
        2
    )
    expense_total = round(
        sum(normalize_amount_for_excel(row.get("支出")) for row in ledger_rows),
        2
    )

    # 明细合计行：只写摘要、支出合计、收入合计；不写 H 列余额。
    total_row = data_start_row + len(ledger_rows)

    if "摘要" in col_map:
        safe_set_cell_value(ws, total_row, col_map["摘要"], "合计")
    if "支出" in col_map:
        safe_set_cell_value(ws, total_row, col_map["支出"], expense_total)
    if "收入" in col_map:
        safe_set_cell_value(ws, total_row, col_map["收入"], income_total)
    if "余额" in col_map:
        safe_set_cell_value(ws, total_row, col_map["余额"], "")
    # 不写余额列数值，保持为空。

    # 空一行后写收入组成
    composition_title_row = total_row + 2
    composition_rows = build_income_composition_rows(ledger_rows)

    summary_col = col_map.get("摘要", 5)
    income_col = col_map.get("收入", 7)

    safe_set_cell_value(ws, composition_title_row, summary_col, "收入组成：")

    for idx, item in enumerate(composition_rows, start=1):
        excel_row = composition_title_row + idx
        safe_set_cell_value(ws, excel_row, summary_col, item["项目"])
        safe_set_cell_value(ws, excel_row, income_col, item["收入"])

    wb.save(output_path)
    return output_path


def generate_ledger_excel(
    matched_df: pd.DataFrame,
    template_path: Path,
    output_dir: Path,
    run_id: str
) -> dict[str, Any]:
    """
    生成台账草稿 Excel，并返回文件信息。
    """
    ledger_rows = build_ledger_rows(matched_df)

    output_path = output_dir / f"党费台账草稿_{run_id}.xlsx"

    write_ledger_rows_to_template(
        ledger_rows=ledger_rows,
        template_path=template_path,
        output_path=output_path
    )

    income_total = 0.0
    expense_total = 0.0

    for row in ledger_rows:
        income_total += normalize_amount_for_excel(row.get("收入"))
        expense_total += normalize_amount_for_excel(row.get("支出"))

    # 为了给接口返回一个财务检查值，这里仍计算 ending_balance；
    # 但不会写入 Excel 台账 H 列。
    ending_balance = round(income_total - expense_total, 2)
    income_composition = build_income_composition_rows(ledger_rows)

    return {
        "ledger_file_name": output_path.name,
        "ledger_file_path": str(output_path),
        "ledger_row_count": len(ledger_rows),
        "ledger_income_total": round(income_total, 2),
        "ledger_expense_total": round(expense_total, 2),
        "ledger_ending_balance": ending_balance,
        "ledger_income_composition": income_composition,
        "ledger_rows_preview": ledger_rows[:20],
    }