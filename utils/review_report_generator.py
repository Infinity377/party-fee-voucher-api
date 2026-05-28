from __future__ import annotations

from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.worksheet import Worksheet


def normalize_records(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """
    将记录转为适合写入 Excel 的普通字典。
    """
    normalized: list[dict[str, Any]] = []

    for item in records:
        normalized.append({
            str(key): "" if value is None else value
            for key, value in item.items()
        })

    return normalized


def autosize_columns(ws: Worksheet) -> None:
    """
    自动调整列宽。
    使用列序号生成列字母，避免 MergedCell 没有 column_letter 属性导致 Pylance 报错。
    """
    from openpyxl.utils import get_column_letter

    for col_idx in range(1, ws.max_column + 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)

        for row_idx in range(1, ws.max_row + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is None:
                continue

            max_length = max(max_length, len(str(value)))

        adjusted_width = min(max(max_length + 2, 12), 80)
        ws.column_dimensions[column_letter].width = adjusted_width


def style_worksheet(ws: Worksheet) -> None:
    """
    简单美化工作表。
    """
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    thin_side = Side(style="thin", color="D9D9D9")
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = border

    if ws.max_row >= 1:
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    autosize_columns(ws)


def build_summary_rows(review_result: dict[str, Any]) -> list[dict[str, Any]]:
    """
    生成复核结论工作表数据。

    新版复核采用“业务实质复核”，不再把上传凭证/台账和基础规则生成结果逐行比较。
    """
    voucher_balance = review_result.get("voucher_balance", {}) or {}

    return [
        {"项目": "复核状态", "结果": review_result.get("review_status", "")},
        {"项目": "是否通过", "结果": "通过" if review_result.get("review_passed") else "不通过"},
        {"项目": "复核异常数量", "结果": review_result.get("review_exception_count", 0)},
        {"项目": "基础规则成功匹配业务数量", "结果": review_result.get("matched_count", 0)},
        {"项目": "原始业务异常数量", "结果": review_result.get("original_business_exception_count", 0)},
        {"项目": "AI补录业务数量", "结果": review_result.get("ai_supplement_count", 0)},
        {"项目": "银行流水覆盖数量", "结果": review_result.get("covered_flow_count", 0)},
        {"项目": "上传凭证业务数量", "结果": review_result.get("voucher_business_count", 0)},
        {"项目": "上传凭证行数", "结果": review_result.get("actual_voucher_row_count", 0)},
        {"项目": "上传台账明细行数", "结果": review_result.get("actual_ledger_row_count", 0)},
        {"项目": "凭证借方合计", "结果": voucher_balance.get("debit_total", 0)},
        {"项目": "凭证贷方合计", "结果": voucher_balance.get("credit_total", 0)},
        {"项目": "借贷平衡检查", "结果": "通过" if voucher_balance.get("balance_check") else "不通过"},
        {"项目": "复核逻辑说明", "结果": "本次采用业务实质复核：核对银行流水、OA流程、业务映射规则、党员状态表、会计科目表与上传凭证/台账之间的业务覆盖关系；AI补录业务单独提示，需人工复核，不按基础规则行数差异列为异常。"},
        {"项目": "复核报告", "结果": review_result.get("review_report", "")},
    ]


def write_sheet(writer: pd.ExcelWriter, sheet_name: str, rows: list[dict[str, Any]]) -> None:
    """
    写入一个工作表。
    """
    if rows:
        df = pd.DataFrame(rows)
    else:
        df = pd.DataFrame([{"提示": "无记录"}])

    df.to_excel(writer, sheet_name=sheet_name, index=False)


def append_review_exception_summary(ws: Worksheet, review_exceptions: list[dict[str, Any]]) -> None:
    """
    在“复核结论”工作表中追加完整异常清单摘要。
    """
    title_fill = PatternFill("solid", fgColor="BDD7EE")
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    thin_side = Side(style="thin", color="D9D9D9")
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    row = ws.max_row + 2

    ws.cell(row=row, column=1, value="四、异常清单摘要（完整明细）")
    ws.cell(row=row, column=1).font = Font(bold=True, size=12)
    ws.cell(row=row, column=1).fill = title_fill
    ws.cell(row=row, column=1).alignment = Alignment(vertical="center", wrap_text=True)
    ws.cell(row=row, column=1).border = border
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    row += 1

    if not review_exceptions:
        ws.cell(row=row, column=1, value="系统复核无异常。AI补录业务如有列示，仅作为待人工复核提示，不计入异常。")
        ws.cell(row=row, column=1).alignment = Alignment(vertical="top", wrap_text=True)
        ws.cell(row=row, column=1).border = border
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        return

    headers = ["序号", "异常类型", "异常说明", "修正建议"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    row += 1

    for idx, item in enumerate(review_exceptions, start=1):
        exception_type = item.get("type", "")
        message = item.get("message", "")
        suggestion = item.get("suggestion", "")

        values = [idx, exception_type, message, suggestion]

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border

        row += 1

    row += 1
    ws.cell(row=row, column=1, value="说明")
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=1).fill = header_fill
    ws.cell(row=row, column=1).alignment = Alignment(vertical="center", wrap_text=True)
    ws.cell(row=row, column=1).border = border

    ws.cell(
        row=row,
        column=2,
        value="以上异常为业务实质复核后的异常，不再包含因AI补录导致的行数差异或后续行号错位。请结合“复核异常清单”工作表逐项核对。"
    )
    ws.cell(row=row, column=2).alignment = Alignment(vertical="top", wrap_text=True)
    ws.cell(row=row, column=2).border = border
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 100
    ws.column_dimensions["D"].width = 70


def append_ai_supplement_summary(ws: Worksheet, ai_supplement_items: list[dict[str, Any]]) -> None:
    """
    在“复核结论”工作表中追加 AI 补录业务提示。
    """
    title_fill = PatternFill("solid", fgColor="E2F0D9")
    header_fill = PatternFill("solid", fgColor="D9EAD3")
    thin_side = Side(style="thin", color="D9D9D9")
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    row = ws.max_row + 2

    ws.cell(row=row, column=1, value="五、AI补录业务提示")
    ws.cell(row=row, column=1).font = Font(bold=True, size=12)
    ws.cell(row=row, column=1).fill = title_fill
    ws.cell(row=row, column=1).alignment = Alignment(vertical="center", wrap_text=True)
    ws.cell(row=row, column=1).border = border
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    row += 1

    if not ai_supplement_items:
        ws.cell(row=row, column=1, value="本次未识别到AI建议补录业务。")
        ws.cell(row=row, column=1).alignment = Alignment(vertical="top", wrap_text=True)
        ws.cell(row=row, column=1).border = border
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        return

    headers = ["序号", "流水序号", "日期", "方向", "对方单位/户名", "金额", "复核提示"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    row += 1

    for idx, item in enumerate(ai_supplement_items, start=1):
        values = [
            idx,
            item.get("flow_index", ""),
            item.get("transaction_date", ""),
            item.get("direction", ""),
            item.get("counterparty", ""),
            item.get("amount", ""),
            item.get("review_prompt", ""),
        ]

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border

        row += 1


def style_review_summary_sheet(ws: Worksheet) -> None:
    """
    对“复核结论”工作表做额外样式优化。
    """
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 100
    ws.column_dimensions["C"].width = 95
    ws.column_dimensions["D"].width = 70
    ws.column_dimensions["E"].width = 36
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 80

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for row_idx in range(1, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 36


def build_ai_supplement_rows(review_result: dict[str, Any]) -> list[dict[str, Any]]:
    """
    生成 AI 补录业务提示工作表数据。
    """
    items = review_result.get("ai_supplement_items", []) or []

    if not items:
        return []

    rows: list[dict[str, Any]] = []

    for item in items:
        rows.append({
            "流水序号": item.get("flow_index", ""),
            "交易日期": item.get("transaction_date", ""),
            "方向": item.get("direction", ""),
            "对方单位/户名": item.get("counterparty", ""),
            "金额": item.get("amount", ""),
            "候选/凭证科目": item.get("subject_code", ""),
            "摘要": item.get("summary", ""),
            "凭证行号": ",".join(str(x) for x in item.get("voucher_rows", [])),
            "复核提示": item.get("review_prompt", ""),
        })

    return rows


def generate_review_report_excel(
    review_result: dict[str, Any],
    output_dir: Path,
    run_id: str
) -> dict[str, Any]:
    """
    生成复核报告 Excel。
    """
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f"党费复核报告_{run_id}.xlsx"

    summary_rows = build_summary_rows(review_result)
    review_exception_rows = normalize_records(review_result.get("review_exceptions", []) or [])
    original_exception_rows = normalize_records(review_result.get("original_business_exceptions", []) or [])
    ai_supplement_rows = normalize_records(build_ai_supplement_rows(review_result))

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        write_sheet(writer, "复核结论", summary_rows)
        write_sheet(writer, "复核异常清单", review_exception_rows)
        write_sheet(writer, "原始业务异常", original_exception_rows)
        write_sheet(writer, "AI补录业务提示", ai_supplement_rows)

    wb = load_workbook(output_path)

    if "复核结论" in wb.sheetnames:
        ws_summary = wb["复核结论"]
        append_review_exception_summary(ws_summary, review_exception_rows)
        append_ai_supplement_summary(ws_summary, review_result.get("ai_supplement_items", []) or [])

    for ws in wb.worksheets:
        style_worksheet(ws)

    if "复核结论" in wb.sheetnames:
        style_review_summary_sheet(wb["复核结论"])

    if "AI补录业务提示" in wb.sheetnames:
        ws_ai = wb["AI补录业务提示"]
        ws_ai.column_dimensions["A"].width = 12
        ws_ai.column_dimensions["B"].width = 16
        ws_ai.column_dimensions["C"].width = 12
        ws_ai.column_dimensions["D"].width = 10
        ws_ai.column_dimensions["E"].width = 36
        ws_ai.column_dimensions["F"].width = 14
        ws_ai.column_dimensions["G"].width = 20
        ws_ai.column_dimensions["H"].width = 80
        ws_ai.column_dimensions["I"].width = 18
        ws_ai.column_dimensions["J"].width = 100

    wb.save(output_path)

    return {
        "review_report_file_name": output_path.name,
        "review_report_file_path": str(output_path),
        "review_report_sheet_count": 4,
    }
