from __future__ import annotations

from pathlib import Path
from typing import Any, cast
import shutil
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from utils.cleaners import normalize_cell


def normalize_amount_for_excel(value: Any) -> float:
    """
    将金额转为 Excel 可写入的 float。
    """
    if value is None or pd.isna(value):
        return 0.0

    if isinstance(value, (int, float)):
        return round(float(value), 2)

    text = str(value).strip().replace(",", "").replace("，", "")
    if text == "":
        return 0.0

    try:
        return round(float(text), 2)
    except ValueError:
        return 0.0


def get_voucher_month_key(voucher_date: str) -> str:
    """
    根据凭证日期生成月份键。
    例如：2026-03-31 -> 2026-03
    """
    voucher_date = normalize_cell(voucher_date)

    try:
        dt = pd.to_datetime(voucher_date)
        return dt.strftime("%Y-%m")
    except Exception:
        if len(voucher_date) >= 7:
            return voucher_date[:7]
        return "unknown"


def get_transaction_date(row: pd.Series) -> str:
    """
    取流水交易日期，用于凭证中的结算日期、核销业务日期。
    优先使用 transaction_date；没有则从 transaction_datetime 中提取日期。
    """
    transaction_date = normalize_cell(row.get("transaction_date"))
    if transaction_date:
        return transaction_date

    transaction_datetime = normalize_cell(row.get("transaction_datetime"))
    if transaction_datetime:
        try:
            return pd.to_datetime(transaction_datetime).strftime("%Y-%m-%d")
        except Exception:
            return transaction_datetime[:10]

    return ""


def build_voucher_rows(
    matched_df: pd.DataFrame,
    maker: str = "",
    book_code: str = "",
    voucher_type: str = ""
) -> list[dict[str, Any]]:
    """
    根据匹配后的业务明细生成凭证行。

    每笔 matched 业务生成两行：
    1. 借方
    2. 贷方

    日期规则：
    1. 制单日期：取交易月份月末 voucher_date；
    2. 结算日期：取流水交易日期 transaction_date；
    3. 核销业务日期：取流水交易日期 transaction_date。

    表头自定义项3规则：
    1. 按凭证制单月份单独编号；
    2. 每个月从 1 开始；
    3. 同一笔业务的借贷两行使用同一个编号。
    """
    rows: list[dict[str, Any]] = []

    maker = maker or "何家俊"
    book_code = book_code or "501-0007"
    voucher_type = voucher_type or "01"

    matched_only = matched_df[matched_df["match_status"] == "matched"].copy()
    matched_only = matched_only.reset_index(drop=True)

    month_group_counter: dict[str, int] = {}

    for _, row in matched_only.iterrows():
        direction = normalize_cell(row.get("direction"))
        amount = normalize_amount_for_excel(row.get("amount"))
        voucher_date = normalize_cell(row.get("voucher_date"))
        settlement_date = get_transaction_date(row)
        summary = normalize_cell(row.get("summary"))
        subject_code = normalize_cell(row.get("subject_code"))

        if amount <= 0:
            continue

        month_key = get_voucher_month_key(voucher_date)
        current_group_no = month_group_counter.get(month_key, 0) + 1
        month_group_counter[month_key] = current_group_no

        common_fields: dict[str, Any] = {
            "核算账簿": book_code,
            "凭证类别": voucher_type,
            "凭证号": "",
            "附单据数": "",
            "制单人": maker,
            "制单日期": voucher_date,
            "审核人": "",
            "审核日期": "",
            "摘要": summary,
            "表头自定义项2": "",
            "表头自定义项3": current_group_no,
            "币种": "CNY",
            "票据号": "",
            "结算日期": settlement_date,
            "结算方式": "",
            "核销号": "",
            "核销业务日期": settlement_date,
        }

        if direction == "收入":
            debit_row = {
                **common_fields,
                "科目编码": "1002",
                "原币借方金额": amount,
                "本币借方金额": amount,
                "原币贷方金额": "",
                "本币贷方金额": "",
            }

            credit_row = {
                **common_fields,
                "科目编码": subject_code,
                "原币借方金额": "",
                "本币借方金额": "",
                "原币贷方金额": amount,
                "本币贷方金额": amount,
            }

            rows.extend([debit_row, credit_row])

        elif direction == "支出":
            debit_row = {
                **common_fields,
                "科目编码": subject_code,
                "原币借方金额": amount,
                "本币借方金额": amount,
                "原币贷方金额": "",
                "本币贷方金额": "",
            }

            credit_row = {
                **common_fields,
                "科目编码": "1002",
                "原币借方金额": "",
                "本币借方金额": "",
                "原币贷方金额": amount,
                "本币贷方金额": amount,
            }

            rows.extend([debit_row, credit_row])

    return rows


def find_header_row(ws: Worksheet) -> int:
    """
    在凭证模板中寻找中文字段行。
    """
    max_scan_row = min(ws.max_row, 10)

    for row_idx in range(1, max_scan_row + 1):
        row_values = [
            str(ws.cell(row=row_idx, column=col_idx).value or "")
            for col_idx in range(1, ws.max_column + 1)
        ]
        row_text = "|".join(row_values)

        if "核算账簿" in row_text and "凭证类别" in row_text and "科目编码" in row_text:
            return row_idx

    return 2


def normalize_template_header(value: Any) -> str:
    """
    模板字段名清洗。
    """
    text = normalize_cell(value)
    text = text.replace("*", "")
    text = text.replace("＊", "")
    text = text.strip()
    return text


def get_template_column_map(ws: Worksheet, header_row: int) -> dict[str, int]:
    """
    获取凭证模板中文字段到列号的映射。
    """
    col_map: dict[str, int] = {}

    for col_idx in range(1, ws.max_column + 1):
        raw_value = ws.cell(row=header_row, column=col_idx).value
        clean_name = normalize_template_header(raw_value)

        if clean_name:
            col_map[clean_name] = col_idx

    return col_map


def clear_old_data(ws: Worksheet, start_row: int) -> None:
    """
    清空模板中的旧数据，保留表头行。
    """
    if ws.max_row >= start_row:
        ws.delete_rows(start_row, ws.max_row - start_row + 1)


def write_voucher_rows_to_template(
    voucher_rows: list[dict[str, Any]],
    template_path: Path,
    output_path: Path
) -> Path:
    """
    将凭证行写入凭证模板，生成新的凭证草稿 Excel。
    """
    if not template_path.exists():
        raise FileNotFoundError(f"凭证模板不存在：{template_path}")

    output_path.parent.mkdir(parents=True, exist_ok=True)

    shutil.copyfile(template_path, output_path)

    wb = load_workbook(output_path)
    ws = cast(Worksheet, wb.active)

    header_row = find_header_row(ws)
    data_start_row = header_row + 1
    col_map = get_template_column_map(ws, header_row)

    clear_old_data(ws, data_start_row)

    for row_offset, voucher_row in enumerate(voucher_rows):
        excel_row = data_start_row + row_offset

        for field_name, value in voucher_row.items():
            col_idx = col_map.get(field_name)

            if col_idx is None:
                continue

            cell_obj = ws.cell(row=excel_row, column=col_idx)
            cell_obj.value = value

    wb.save(output_path)
    return output_path


def generate_voucher_excel(
    matched_df: pd.DataFrame,
    template_path: Path,
    output_dir: Path,
    run_id: str,
    maker: str = "",
    book_code: str = "",
    voucher_type: str = ""
) -> dict[str, Any]:
    """
    生成凭证草稿 Excel，并返回文件信息和预览数据。
    """
    voucher_rows = build_voucher_rows(
        matched_df=matched_df,
        maker=maker,
        book_code=book_code,
        voucher_type=voucher_type
    )

    output_path = output_dir / f"党费凭证草稿_{run_id}.xlsx"

    write_voucher_rows_to_template(
        voucher_rows=voucher_rows,
        template_path=template_path,
        output_path=output_path
    )

    debit_total = 0.0
    credit_total = 0.0

    for row in voucher_rows:
        debit_total += normalize_amount_for_excel(row.get("本币借方金额"))
        credit_total += normalize_amount_for_excel(row.get("本币贷方金额"))

    return {
        "voucher_file_name": output_path.name,
        "voucher_file_path": str(output_path),
        "voucher_row_count": len(voucher_rows),
        "voucher_business_count": int(len(voucher_rows) / 2),
        "voucher_debit_total": round(debit_total, 2),
        "voucher_credit_total": round(credit_total, 2),
        "voucher_balance_check": abs(debit_total - credit_total) < 0.005,
        "voucher_rows_preview": voucher_rows[:20],
    }